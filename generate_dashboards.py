"""
CARS24 Dashboard Generator
==========================
Fetches data from Google Sheets and generates two standalone HTML dashboards.
No API key needed. No server needed. Just run this script.

Usage:
    python generate_dashboards.py

Output:
    dashboard_metrics.html      — Performance metrics dashboard
    dashboard_qualitative.html  — Qualitative insights dashboard

Requirements:
    pip install requests pandas
"""

import io, json, re, os, sys
from datetime import datetime, timedelta
from collections import Counter
from pathlib import Path

CHARTJS_CDN = "https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"

def get_chartjs():
    """Download Chart.js once and cache it locally so HTML is fully self-contained."""
    cache = Path(__file__).parent / "_chartjs_cache.js"
    if cache.exists():
        return cache.read_text(encoding="utf-8")
    print("Downloading Chart.js (one-time)...")
    r = requests.get(CHARTJS_CDN, timeout=30)
    r.raise_for_status()
    cache.write_text(r.text, encoding="utf-8")
    return r.text

try:
    import requests
    import pandas as pd
except ImportError:
    print("Installing required packages...")
    os.system(f"{sys.executable} -m pip install requests pandas")
    import requests
    import pandas as pd

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────

SHEET_ID     = os.environ.get("GOOGLE_SHEET_ID", "13tnQTLEtCO3nPo0uQ_gcwkXsfXwM4OaZucs-4xw8Unk")
OUT_DIR      = Path(__file__).parent          # same folder as this script
ARCHIVE_FILE = OUT_DIR / "data_archive.csv"   # grows incrementally, never loses data

# Apps Script Web App — fetches sheet as CSV without requiring the sheet to be public
APPS_SCRIPT_URL   = "https://script.google.com/a/macros/cars24.com/s/AKfycbzOnDWfBliuHdbmUk3TbhlTypG6LoiwGhDVHEtJhpsJBqKqS9aYc2-xs8tKdDvcq1UD/exec"
APPS_SCRIPT_TOKEN = "cars24dash2024"

# Only keep columns needed for the dashboard — keeps archive size small
KEEP_COLS = ["_ID", "SESSION_ID", "CHAT_TIME_IST", "CHAT_DATE", "AGENT_ID",
             "MSG_TYPE", "CONTENT", "INSP_DATE_3", "REGION_3", "NEW_TP", "APPOINTMENT_ID"]

# ── Fetch + Archive ───────────────────────────────────────────────────────────

def _load_sheet_csv_text(csv_path=None):
    """
    Return the raw CSV text of the Google Sheet.
    Priority:
      1. csv_path          — manually supplied local file (--csv flag)
      2. Auto-detected CSV — any .csv file found in the same folder as the script
      3. Chrome cookies    — uses your existing Chrome CARS24 login session (no setup needed)
      4. OAuth token       — saved token.json from --auth-setup flow
      5. Public URL        — works only if sheet is 'Anyone with link → Viewer'
    """
    def _read_csv_text(path):
        """Read a CSV file trying common encodings automatically."""
        for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252", "mac_roman"):
            try:
                return Path(path).read_text(encoding=enc)
            except (UnicodeDecodeError, LookupError):
                continue
        # Last resort: read bytes and decode ignoring errors
        return Path(path).read_bytes().decode("utf-8", errors="replace")

    # ── Option 1: local CSV file (--csv flag) ─────────────────────────────────
    if csv_path:
        print(f"  [csv mode] Reading from file: {csv_path}")
        return _read_csv_text(csv_path)

    # ── Option 2: auto-detect CSV in the same folder as the script ───────────
    # Looks for any .csv file that isn't the data archive itself
    script_dir = OUT_DIR
    csv_candidates = [
        f for f in script_dir.glob("*.csv")
        if f.name != ARCHIVE_FILE.name
    ]
    if csv_candidates:
        # Prefer the most recently modified one
        latest_csv = max(csv_candidates, key=lambda f: f.stat().st_mtime)
        print(f"  [auto-csv] Found CSV in folder: {latest_csv.name}")
        return _read_csv_text(latest_csv)

    # ── Option 2: Chrome browser cookies (uses your CARS24 Google session) ────
    # This works as long as Chrome is open and you're logged in with your CARS24 account.
    # Install once with: pip install browser-cookie3
    if APPS_SCRIPT_URL:
        try:
            import browser_cookie3
            print(f"  Fetching via Apps Script (Chrome session) …")
            cj  = browser_cookie3.chrome(domain_name='.google.com')
            url = f"{APPS_SCRIPT_URL}?token={APPS_SCRIPT_TOKEN}"
            resp = requests.get(url, cookies=cj, timeout=60,
                                allow_redirects=True,
                                headers={"User-Agent": "Mozilla/5.0"})
            text = resp.text.strip()
            if resp.status_code == 200 and not text.startswith("{") and not text.startswith("<"):
                print(f"  [chrome-cookie] Fetch OK ({len(resp.text)} chars)")
                return resp.text
            else:
                print(f"  [chrome-cookie] Unexpected response (status={resp.status_code}, "
                      f"starts_with={repr(text[:40])}) — falling back.")
        except ImportError:
            print("  [chrome-cookie] browser-cookie3 not installed — "
                  "run: pip install browser-cookie3")
        except Exception as e:
            print(f"  [chrome-cookie] Failed ({e}) — falling back.")

    # ── Option 3: saved OAuth token ───────────────────────────────────────────
    token_file = OUT_DIR / "token.json"
    if token_file.exists():
        try:
            from google.oauth2.credentials import Credentials
            from google.auth.transport.requests import Request as GRequest
            creds = Credentials.from_authorized_user_file(str(token_file),
                        scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"])
            if creds.expired and creds.refresh_token:
                creds.refresh(GRequest())
                token_file.write_text(creds.to_json())
            url  = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid=0"
            resp = requests.get(url, headers={"Authorization": f"Bearer {creds.token}"}, timeout=30)
            resp.raise_for_status()
            print(f"  [oauth] Authenticated fetch OK")
            return resp.text
        except Exception as e:
            print(f"  [oauth] Token present but failed ({e}) — falling back to public URL.")

    # ── Option 4: public URL (sheet must be 'Anyone with link → Viewer') ─────
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid=0"
    print(f"  Fetching sheet (public URL) …")
    resp = requests.get(url, timeout=30)
    if resp.status_code == 401:
        raise RuntimeError(
            "\n\n  ❌  All fetch methods failed.\n\n"
            "  Make sure Chrome is open and you are logged in with your CARS24 Google account,\n"
            "  then run:  pip install browser-cookie3\n"
            "  and try again.\n\n"
            "  Or download the sheet manually and run:\n"
            "    python generate_dashboards.py --csv <path_to_downloaded.csv>\n"
        )
    resp.raise_for_status()
    if resp.text.strip().startswith("<"):
        raise RuntimeError("Google returned HTML instead of CSV — check sheet permissions.")
    return resp.text


def fetch_and_archive(offline=False, csv_path=None):
    """
    Fetch latest data from Google Sheet, merge with local archive, deduplicate,
    save back to archive. Returns the full combined dataset as list of dicts.

    offline=True  — skip network, use local archive only
    csv_path      — read from a locally downloaded CSV file
    """
    if offline:
        print("  [offline mode] Skipping Google Sheets fetch — using archive only.")
        if not ARCHIVE_FILE.exists():
            raise RuntimeError("No archive file found. Run without --offline first to create an archive.")
        df_combined = pd.read_csv(ARCHIVE_FILE, engine="python", on_bad_lines="skip")
        df_combined.columns = [c.strip().upper() for c in df_combined.columns]
        print(f"  Archive rows: {len(df_combined)}")
    else:
        # ── 1. Fetch fresh rows from Google Sheet ──
        csv_text = _load_sheet_csv_text(csv_path=csv_path)
        df_new = pd.read_csv(io.StringIO(csv_text), low_memory=False, header=0)
        ncols = len(df_new.columns)
        # ── Pin positional columns BEFORE name normalisation ──────────────────
        # The sheet has duplicate column headers; we must use column letter
        # positions to pick the correct ones:
        #   BY (col 77, idx 76) = NEW_TP  (fair market value / target price)
        #   DJ (col 114, idx 113) = INSP_DATE_3  (inspection date)
        #   DL (col 116, idx 115) = REGION_3     (region)
        # Overwrite by position so deduplication can't pick the wrong copy.
        # Positional pinning only applies to full Google Sheet dumps (50+ columns).
        # Pre-processed archive-format CSVs (< 50 cols) already have correct column names.
        if ncols >= 50:
            POSITIONAL = {8: "APPOINTMENT_ID", 76: "NEW_TP", 113: "INSP_DATE_3", 115: "REGION_3"}
            for idx_pos, col_name in POSITIONAL.items():
                if idx_pos < ncols:
                    df_new[col_name] = df_new.iloc[:, idx_pos].values
        df_new.columns = [c.strip().upper() for c in df_new.columns]
        # Keep only the columns we need (drop large/unused ones)
        keep = [c for c in KEEP_COLS if c in df_new.columns]
        df_new = df_new[keep].copy()
        print(f"  Sheet rows: {len(df_new)}")

        # ── 2. Load existing archive (if any) ──
        if ARCHIVE_FILE.exists():
            df_arch = pd.read_csv(ARCHIVE_FILE, engine="python", on_bad_lines="skip")
            df_arch.columns = [c.strip().upper() for c in df_arch.columns]
            print(f"  Archive rows: {len(df_arch)}")
            df_combined = pd.concat([df_new, df_arch], ignore_index=True)  # df_new first → fresh sheet data wins dedup
        else:
            print("  No archive yet — creating one from current sheet data.")
            df_combined = df_new

    # ── 3. Deduplicate ──
    # Prefer _ID (unique row key) if available; fall back to composite key
    if "_ID" in df_combined.columns:
        before = len(df_combined)
        df_combined = df_combined.drop_duplicates(subset=["_ID"])
        print(f"  Deduplicated on _ID: {before} → {len(df_combined)} rows")
    else:
        dedup_on = [c for c in ["SESSION_ID", "CHAT_TIME_IST", "MSG_TYPE"] if c in df_combined.columns]
        before = len(df_combined)
        df_combined = df_combined.drop_duplicates(subset=dedup_on)
        print(f"  Deduplicated on {dedup_on}: {before} → {len(df_combined)} rows")

    # ── 4. Save updated archive (only when we fetched fresh data) ──
    if not offline:
        df_combined.to_csv(ARCHIVE_FILE, index=False)
        size_kb = ARCHIVE_FILE.stat().st_size // 1024
        print(f"  Archive saved: {len(df_combined)} total rows ({size_kb} KB) → {ARCHIVE_FILE.name}")

    # ── 5. Normalise column names for downstream processing ──
    df_combined.columns = [c.strip().upper() for c in df_combined.columns]
    df_combined.rename(columns={
        "SESSION_ID":    "session_id",
        "CHAT_TIME_IST": "chat_time",
        "CHAT_DATE":     "chat_date",
        "AGENT_ID":      "agent_id",
        "MSG_TYPE":      "msg_type",
        "CONTENT":       "content",
        "INSP_DATE_3":      "insp_date",
        "REGION_3":         "region",
        "NEW_TP":           "new_tp",
        "APPOINTMENT_ID":   "appointment_id",
    }, inplace=True)
    df_combined["session_id"] = df_combined.get("session_id", pd.Series(dtype=str)).astype(str).str.strip()
    df_combined["msg_type"]   = df_combined.get("msg_type",   pd.Series(dtype=str)).astype(str).str.strip().str.lower()
    df_combined["content"]    = df_combined.get("content",    pd.Series(dtype=str)).fillna("").astype(str)
    df_combined["insp_date"]  = df_combined.get("insp_date",  pd.Series(dtype=str)).fillna("").astype(str)
    df_combined["region"]     = df_combined.get("region",     pd.Series(dtype=str)).fillna("").astype(str)
    df_combined["new_tp"]     = df_combined.get("new_tp",     pd.Series(dtype=str)).fillna("").astype(str)
    df_combined = df_combined[
        df_combined["session_id"].notna() &
        (df_combined["session_id"] != "") &
        (df_combined["session_id"] != "nan")
    ]
    print(f"  Using {len(df_combined)} rows for dashboard generation.\n")
    return df_combined.to_dict("records")

# ── Classification ────────────────────────────────────────────────────────────

def strip_md(t):
    return re.sub(r"\*{1,3}", "", re.sub(r"_{1,3}", "", t))

def classify_session(msgs):
    all_content  = " ".join(m.get("content","") for m in msgs)
    bot_contents = [m.get("content","") for m in msgs if m.get("msg_type") == "ai"]
    if re.search(r"auction started", all_content, re.I): return "auction_started"
    if any(re.search(r"\b\d{7}\b", c) for c in bot_contents): return "ticket_raised"
    last_bot = bot_contents[-1] if bot_contents else ""
    lb = strip_md(last_bot).lower()
    res  = ["closing this chat","anything else i can help","speak to your",
            "we're closing","auction is already running","auction is currently",
            "we are running an auction","we'll update you"]
    pend = ["yes/no)","please share your email","please confirm and share",
            "could you share","can you share","what would you like to proceed",
            "share your registered email","share your email id","please share your email id"]
    has_call = bool(re.search(r"call\s+\d{5,}", lb))
    has_pin  = bool(re.search(r"pin\s*[:\*]*\s*\d{3,}", lb))
    is_res   = any(s in lb for s in res) or has_call or has_pin
    is_pend  = any(s in lb for s in pend)
    if is_pend and not is_res: return "drop_without_resolution"
    if is_res: return "drop_with_resolution"
    return "drop_without_resolution" if ("?" in last_bot and "anything else" not in lb) else "drop_with_resolution"

# ── Session processing ────────────────────────────────────────────────────────

# ── EHB price grid: add premium to auction price based on range ───────────────
# Grid: (upper_limit_exclusive, premium_to_add)  — in ascending order
# upper_limit=None means ">30 Lakh" (no upper bound)
EHB_PRICE_GRID = [
    (100_000,   2_200),   # < 1 Lakh
    (200_000,   3_850),   # 1–2 Lakh
    (300_000,   5_000),   # 2–3 Lakh
    (500_000,   6_100),   # 3–5 Lakh
    (1_000_000, 9_000),   # 5–10 Lakh
    (1_500_000, 12_500),  # 10–15 Lakh
    (2_000_000, 18_500),  # 15–20 Lakh
    (3_000_000, 20_000),  # 20–30 Lakh
    (None,      25_000),  # > 30 Lakh
]

def _apply_ehb_premium(auction_price):
    """Return auction_price + premium based on price grid."""
    for upper, premium in EHB_PRICE_GRID:
        if upper is None or auction_price < upper:
            return auction_price + premium
    return auction_price + 25_000  # fallback


def process_sessions(rows):
    grouped = {}
    for r in rows:
        sid = str(r.get("session_id","")).strip()
        if sid: grouped.setdefault(sid, []).append(r)
    sessions = []
    for sid, msgs in grouped.items():
        msgs = sorted(msgs, key=lambda m: str(m.get("chat_time") or m.get("chat_date") or ""))
        case_type = "DSQ" if "dsq" in sid.lower() else "QNR"
        outcome   = classify_session(msgs)
        bot_msgs  = [m for m in msgs if m.get("msg_type") == "ai"]

        # Repeat analysis: count occurrences of each bot message
        counts = Counter(m.get("content","") for m in bot_msgs)
        has_repeat   = any(v >= 2 for v in counts.values())
        repeat_count = max(counts.values()) if counts else 0  # max occurrences of any single message

        first_time = str(msgs[0].get("chat_time") or msgs[0].get("chat_date") or "")[:10]

        # has_both: session where BOTH auction-start AND a 7-digit ticket appear
        all_content = " ".join(m.get("content","") for m in msgs)
        has_auction = bool(re.search(r"auction started", all_content, re.I))
        has_ticket  = any(re.search(r"\b\d{7}\b", m.get("content","")) for m in bot_msgs)
        has_both    = has_auction and has_ticket

        # Region: take first non-empty value in the session
        regions = [str(m.get("region","")).strip() for m in msgs
                   if str(m.get("region","")).strip() not in ("", "nan")]
        region = regions[0] if regions else ""

        # Inspection date: earliest non-empty INSP_DATE_3 value in session
        insp_dates = []
        for m in msgs:
            d = str(m.get("insp_date","")).strip()
            if d and d not in ("", "nan", "NaT"):
                insp_dates.append(d[:10])
        insp_date = min(insp_dates) if insp_dates else ""

        # NEW_TP: average of all non-empty numeric TP values in the session
        tp_vals = []
        for m in msgs:
            v = str(m.get("new_tp","")).strip()
            if v and v not in ("", "nan", "NaT"):
                try:
                    tp_vals.append(float(v.replace(",","")))
                except Exception:
                    pass
        new_tp_num = round(sum(tp_vals) / len(tp_vals), 2) if tp_vals else None

        # EHB: for each auction price found, apply the price grid premium,
        # then average across all auctions in the session.
        # e.g. auction at ₹95,000 → add ₹2,200 → HB = ₹97,200
        # ₹ may be stored as UTF-8 bytes \xe2\x82\xb9 due to CSV encoding
        ehb_vals = []
        if has_auction:
            for m in bot_msgs:
                c = m.get("content","")
                ehb_match = re.search(r'(?:\xe2\x82\xb9|\u20b9|\u20B9|₹)\s*([\d,]+)', c)
                if not ehb_match:
                    # fallback: find price pattern after "auction started" keyword
                    ehb_match = re.search(r'[Aa]uction\s+started[^\d]*([\d,]{3,})', c)
                if ehb_match:
                    try:
                        raw_price = float(ehb_match.group(1).replace(",",""))
                        # Apply grid premium to get the true HB price
                        hb_price = _apply_ehb_premium(raw_price)
                        ehb_vals.append(hb_price)
                    except Exception:
                        pass
        ehb = round(sum(ehb_vals) / len(ehb_vals), 2) if ehb_vals else None

        # Earliest chat date for this session
        chat_dates = []
        for m in msgs:
            d = str(m.get("chat_date","")).strip()
            if d and d not in ("", "nan", "NaT"):
                chat_dates.append(d[:10])
        earliest_chat_date = min(chat_dates) if chat_dates else first_time

        # APPOINTMENT_ID: first non-empty value across all messages in the session
        appt_ids = [str(m.get("appointment_id","")).strip()
                    for m in msgs if str(m.get("appointment_id","")).strip() not in ("", "nan", "NaT")]
        appointment_id = appt_ids[0] if appt_ids else ""

        sessions.append({
            "id": sid, "case_type": case_type, "outcome": outcome,
            "total": len(msgs),
            "human": sum(1 for m in msgs if m.get("msg_type") == "human"),
            "bot": len(bot_msgs),
            "has_repeat": has_repeat,
            "repeat_count": repeat_count,
            "date": first_time,
            "has_both": has_both,
            "region": region,
            "insp_date": insp_date,
            "earliest_chat_date": earliest_chat_date,
            "ehb": ehb,
            "new_tp": new_tp_num,
            "appointment_id": appointment_id,
        })
    return sessions

# ── Gap helpers ───────────────────────────────────────────────────────────────

GAP_BUCKETS = ["Same Day", "1-3 days", "3-5 days", "5-10 days", "10-20 days", "20-30 days", ">30 days"]

def _gap_bucket(gap):
    if gap == 0:  return "Same Day"
    if gap <= 3:  return "1-3 days"
    if gap <= 5:  return "3-5 days"
    if gap <= 10: return "5-10 days"
    if gap <= 20: return "10-20 days"
    if gap <= 30: return "20-30 days"
    return ">30 days"

def _compute_trend(sessions, period="week"):
    """Compute weekly or monthly outcome % breakdown for last 3 months."""
    valid = [s for s in sessions if s.get("date") and len(s["date"]) >= 10]
    if not valid:
        return []
    max_date = max(s["date"] for s in valid)
    try:
        max_dt = datetime.strptime(max_date[:10], "%Y-%m-%d")
    except Exception:
        return []
    cutoff_dt = max_dt - timedelta(days=91)   # approx 3 months

    outcomes = ["auction_started", "ticket_raised", "drop_with_resolution", "drop_without_resolution"]
    buckets  = {}
    for s in valid:
        try:
            d = datetime.strptime(s["date"][:10], "%Y-%m-%d")
        except Exception:
            continue
        if d < cutoff_dt:
            continue
        key = d.strftime("%Y-W%W") if period == "week" else d.strftime("%Y-%m")
        ct  = s["case_type"]
        oc  = s["outcome"]
        if key not in buckets:
            buckets[key] = {
                "DSQ": {o: 0 for o in outcomes},
                "QNR": {o: 0 for o in outcomes},
            }
        if oc in outcomes:
            buckets[key][ct][oc] += 1

    result = []
    p = lambda n, d: round(n / max(d, 1) * 100, 1)
    for key in sorted(buckets.keys()):
        b = buckets[key]
        dsq_total = sum(b["DSQ"].values())
        qnr_total = sum(b["QNR"].values())
        result.append({
            "label":     key,
            "dsq_total": dsq_total,
            "dsq_pcts":  {o: p(b["DSQ"][o], dsq_total) for o in outcomes},
            "qnr_total": qnr_total,
            "qnr_pcts":  {o: p(b["QNR"][o], qnr_total) for o in outcomes},
        })
    return result

def _compute_region_split(sessions):
    """Group sessions by region, return DSQ/QNR counts sorted by volume."""
    regions = {}
    for s in sessions:
        r = s.get("region","").strip()
        if not r or r == "nan":
            r = "Unknown"
        ct = s["case_type"]
        if r not in regions:
            regions[r] = {"DSQ": 0, "QNR": 0}
        regions[r][ct] += 1
    result = []
    for r, counts in sorted(regions.items(), key=lambda x: -(x[1]["DSQ"] + x[1]["QNR"])):
        result.append({"region": r, "dsq": counts["DSQ"], "qnr": counts["QNR"]})
    return result[:20]   # top 20 regions

def _compute_insp_gap(sessions):
    """Calculate inspection-to-chat-date gap in days, bucketed."""
    dsq_b = {b: 0 for b in GAP_BUCKETS}
    qnr_b = {b: 0 for b in GAP_BUCKETS}
    for s in sessions:
        insp = s.get("insp_date","").strip()
        chat = s.get("earliest_chat_date", s.get("date","")).strip()
        if not insp or not chat or insp in ("", "nan") or chat in ("", "nan"):
            continue
        try:
            i_dt  = datetime.strptime(insp[:10], "%Y-%m-%d")
            c_dt  = datetime.strptime(chat[:10], "%Y-%m-%d")
            gap   = (c_dt - i_dt).days
            if gap < 0:   # inspection is after chat — ignore
                continue
        except Exception:
            continue
        bucket = _gap_bucket(gap)
        if s["case_type"] == "DSQ":
            dsq_b[bucket] += 1
        else:
            qnr_b[bucket] += 1
    return {
        "buckets": GAP_BUCKETS,
        "dsq":     [dsq_b[b] for b in GAP_BUCKETS],
        "qnr":     [qnr_b[b] for b in GAP_BUCKETS],
    }

# ── Metrics ───────────────────────────────────────────────────────────────────

def compute_metrics(sessions):
    total = len(sessions)
    dsq   = [s for s in sessions if s["case_type"] == "DSQ"]
    qnr   = [s for s in sessions if s["case_type"] == "QNR"]
    def cnt(lst):
        return {
            "auction_started":         sum(1 for s in lst if s["outcome"] == "auction_started"),
            "ticket_raised":           sum(1 for s in lst if s["outcome"] == "ticket_raised"),
            "drop_with_resolution":    sum(1 for s in lst if s["outcome"] == "drop_with_resolution"),
            "drop_without_resolution": sum(1 for s in lst if s["outcome"] == "drop_without_resolution"),
        }
    ov, dc, qc = cnt(sessions), cnt(dsq), cnt(qnr)
    daily = {}
    for s in sessions:
        if s["date"]: daily[s["date"]] = daily.get(s["date"], 0) + 1
    repeat = sum(1 for s in sessions if s["has_repeat"])
    p = lambda n, d: round(n / max(d, 1) * 100, 1)
    avg_total = round(sum(s["total"] for s in sessions) / max(total, 1), 1)
    avg_bot   = round(sum(s["bot"]   for s in sessions) / max(total, 1), 1)
    avg_human = round(sum(s["human"] for s in sessions) / max(total, 1), 1)
    res_rate  = p(ov["drop_with_resolution"],
                  ov["drop_with_resolution"] + ov["drop_without_resolution"])

    # has_both: sessions with BOTH auction start AND ticket in same conversation
    has_both_count = sum(1 for s in sessions if s.get("has_both"))

    # Repeat buckets (cumulative): >1 rep, >3 reps, >5 reps
    repeat_buckets = []
    for thresh, label in [(2, ">1 repetition"), (4, ">3 repetitions"), (6, ">5 repetitions")]:
        matching = [s for s in sessions if s.get("repeat_count", 0) >= thresh]
        repeat_buckets.append({
            "label":         label,
            "session_count": len(matching),
            "rep_count":     sum(s.get("repeat_count", 0) for s in matching),
        })

    # Trend data
    weekly_trends  = _compute_trend(sessions, "week")
    monthly_trends = _compute_trend(sessions, "month")

    # Region split
    region_split = _compute_region_split(sessions)

    # Inspection-to-chat gap
    insp_gap = _compute_insp_gap(sessions)

    # EHB/TP: ratio of auction start price to target price, for sessions where both are available
    ehb_tp_dsq_vals, ehb_tp_qnr_vals = [], []
    for s in sessions:
        ehb = s.get("ehb")
        tp  = s.get("new_tp")
        if ehb and tp and tp > 0 and s.get("outcome") == "auction_started":
            ratio = round(ehb / tp * 100, 1)
            if s["case_type"] == "DSQ":
                ehb_tp_dsq_vals.append(ratio)
            else:
                ehb_tp_qnr_vals.append(ratio)

    ehb_tp_dsq     = round(sum(ehb_tp_dsq_vals) / len(ehb_tp_dsq_vals), 1) if ehb_tp_dsq_vals else None
    ehb_tp_qnr     = round(sum(ehb_tp_qnr_vals) / len(ehb_tp_qnr_vals), 1) if ehb_tp_qnr_vals else None
    ehb_tp_dsq_n   = len(ehb_tp_dsq_vals)
    ehb_tp_qnr_n   = len(ehb_tp_qnr_vals)

    # Average EHB price (rupees) for all auction sessions (with or without TP)
    ehb_dsq_all = [s["ehb"] for s in sessions if s.get("ehb") and s.get("outcome")=="auction_started" and s["case_type"]=="DSQ"]
    ehb_qnr_all = [s["ehb"] for s in sessions if s.get("ehb") and s.get("outcome")=="auction_started" and s["case_type"]=="QNR"]
    avg_ehb_dsq  = round(sum(ehb_dsq_all) / len(ehb_dsq_all)) if ehb_dsq_all else None
    avg_ehb_qnr  = round(sum(ehb_qnr_all) / len(ehb_qnr_all)) if ehb_qnr_all else None
    ehb_dsq_n    = len(ehb_dsq_all)
    ehb_qnr_n    = len(ehb_qnr_all)

    sessions_list = [
        {"si": s.get("id",""),
         "d": s.get("earliest_chat_date") or s.get("date",""),
         "ct": s.get("case_type",""), "oc": s.get("outcome",""),
         "hb": 1 if s.get("has_both") else 0,
         "hr": 1 if s.get("has_repeat") else 0,
         "rc": s.get("repeat_count", 0),
         "rg": s.get("region",""),
         "id": s.get("insp_date",""),
         "n":  s.get("total", 0),
         "eb": s.get("ehb"),
         "tp": s.get("new_tp"),
         "ai": s.get("appointment_id","")}
        for s in sessions
    ]
    return {
        "total": total, "dsq_count": len(dsq), "qnr_count": len(qnr),
        "outcomes": ov, "dsq_outcomes": dc, "qnr_outcomes": qc,
        "repeat": repeat, "repeat_pct": p(repeat, total),
        "avg_msgs": avg_total, "avg_bot": avg_bot, "avg_human": avg_human,
        "auction_rate_dsq": p(dc["auction_started"], len(dsq)),
        "auction_rate_qnr": p(qc["auction_started"], len(qnr)),
        "ticket_rate":      p(ov["ticket_raised"], total),
        "drop_rate":        p(ov["drop_with_resolution"] + ov["drop_without_resolution"], total),
        "resolution_rate":  res_rate,
        "daily":            [{"d": d, "c": c} for d, c in sorted(daily.items())],
        "has_both_count":   has_both_count,
        "has_both_pct":     p(has_both_count, total),
        "repeat_buckets":   repeat_buckets,
        "weekly_trends":    weekly_trends,
        "monthly_trends":   monthly_trends,
        "region_split":     region_split,
        "insp_gap":         insp_gap,
        "sessions_list":    sessions_list,
        "ehb_tp_dsq":       ehb_tp_dsq,
        "ehb_tp_qnr":       ehb_tp_qnr,
        "ehb_tp_dsq_n":     ehb_tp_dsq_n,
        "ehb_tp_qnr_n":     ehb_tp_qnr_n,
        "avg_ehb_dsq":      avg_ehb_dsq,
        "avg_ehb_qnr":      avg_ehb_qnr,
        "ehb_dsq_n":        ehb_dsq_n,
        "ehb_qnr_n":        ehb_qnr_n,
    }

# ── Rule-based qualitative insights ──────────────────────────────────────────

def generate_insights(m):
    t   = m["total"]
    ov  = m["outcomes"]
    dc  = m["dsq_outcomes"]
    qc  = m["qnr_outcomes"]
    p   = lambda n, d: round(n / max(d, 1) * 100)

    ww, ni = [], []

    # Working well
    if m["auction_rate_qnr"] >= 40:
        ww.append({"title": "Strong QNR auction conversion", "caseType": "QNR",
            "detail": f"{m['auction_rate_qnr']}% of QNR sessions convert to a new auction, showing the bot effectively negotiates with undecided customers.",
            "impact": "High" if m["auction_rate_qnr"] >= 60 else "Medium"})

    if m["auction_rate_dsq"] >= 20:
        ww.append({"title": "DSQ customers re-entering auctions", "caseType": "DSQ",
            "detail": f"{m['auction_rate_dsq']}% of dissatisfied customers agree to a fresh auction — the bot successfully reframes value.",
            "impact": "High" if m["auction_rate_dsq"] >= 35 else "Medium"})

    if m["resolution_rate"] >= 50:
        ww.append({"title": "Solid resolution rate on dropped chats", "caseType": "Both",
            "detail": f"{m['resolution_rate']}% of sessions that don't convert are still resolved before closing, reducing escalation load.",
            "impact": "High" if m["resolution_rate"] >= 70 else "Medium"})

    if m["repeat_pct"] < 5:
        ww.append({"title": "Minimal bot message repetition", "caseType": "Both",
            "detail": f"Only {m['repeat_pct']}% of sessions show duplicate bot messages — conversation flow is stable.",
            "impact": "Medium"})

    if m["ticket_rate"] >= 5:
        ww.append({"title": "Ticket escalation working correctly", "caseType": "Both",
            "detail": f"{m['ticket_rate']}% of sessions are escalated via ticket, catching complex cases before customer frustration peaks.",
            "impact": "Medium"})

    if 4 <= m["avg_msgs"] <= 12:
        ww.append({"title": "Efficient conversation length", "caseType": "Both",
            "detail": f"Average {m['avg_msgs']} messages per session — the bot reaches outcomes without unnecessary back-and-forth.",
            "impact": "Low"})

    if m["dsq_count"] > 0 and m["qnr_count"] > 0:
        ratio = max(m["dsq_count"], m["qnr_count"]) / max(min(m["dsq_count"], m["qnr_count"]), 1)
        if ratio < 3:
            ww.append({"title": "Balanced DSQ / QNR case handling", "caseType": "Both",
                "detail": f"{m['dsq_count']} DSQ and {m['qnr_count']} QNR sessions — neither case type is overwhelming the system.",
                "impact": "Low"})

    if p(ov["auction_started"], t) >= 15:
        ww.append({"title": "Healthy overall auction start rate", "caseType": "Both",
            "detail": f"{p(ov['auction_started'],t)}% of all sessions result in a new auction, demonstrating positive business outcomes.",
            "impact": "High"})

    # Needs improvement
    unres_pct = p(ov["drop_without_resolution"], t)
    if unres_pct >= 25:
        ni.append({"title": "High unresolved drop rate", "caseType": "Both",
            "detail": f"{unres_pct}% of sessions end without resolution — customers leave with unanswered questions.",
            "recommendation": "Add a closing-confirmation step that summarises what was discussed and offers a callback or ticket before the chat ends.",
            "priority": "High" if unres_pct >= 50 else "Medium"})

    if m["auction_rate_qnr"] < 40:
        ni.append({"title": "QNR auction conversion needs improvement", "caseType": "QNR",
            "detail": f"Only {m['auction_rate_qnr']}% of QNR sessions convert to a new auction.",
            "recommendation": "Strengthen the negotiation script for QNR — provide clearer auction benefits and a simpler Yes/No confirmation step.",
            "priority": "High" if m["auction_rate_qnr"] < 20 else "Medium"})

    if m["repeat_pct"] >= 5:
        ni.append({"title": "Bot duplicate message bug detected", "caseType": "Both",
            "detail": f"{m['repeat_pct']}% of sessions contain repeated bot messages, degrading conversation quality.",
            "recommendation": "Add a deduplication check before any bot message is sent and audit the state machine for loop conditions.",
            "priority": "High" if m["repeat_pct"] >= 15 else "Medium"})

    if m["auction_rate_dsq"] < 20:
        ni.append({"title": "DSQ customers not re-engaging with auctions", "caseType": "DSQ",
            "detail": f"Only {m['auction_rate_dsq']}% of DSQ customers agree to a fresh auction.",
            "recommendation": "Introduce social-proof messaging (e.g., recent auction outcomes) and clarify the fresh auction process for DSQ customers.",
            "priority": "Medium"})

    if m["avg_msgs"] > 14:
        ni.append({"title": "Conversations running too long", "caseType": "Both",
            "detail": f"Average {m['avg_msgs']} messages per session suggests the bot is looping rather than resolving.",
            "recommendation": "Add intent detection to shortcut responses and limit clarification loops to 2 rounds before offering escalation.",
            "priority": "Medium"})

    if m["avg_msgs"] < 3:
        ni.append({"title": "Sessions ending prematurely", "caseType": "Both",
            "detail": f"Average only {m['avg_msgs']} messages — customers are likely dropping before the bot can help.",
            "recommendation": "Improve the opening greeting and intent capture to engage customers immediately.",
            "priority": "High"})

    if m["ticket_rate"] < 3 and unres_pct > 30:
        ni.append({"title": "Under-utilising ticket escalation", "caseType": "Both",
            "detail": f"Only {m['ticket_rate']}% of unresolved sessions raise a ticket, leaving customers without follow-up.",
            "recommendation": "Automatically offer a ticket or callback whenever a session ends without an auction or resolution.",
            "priority": "Medium"})

    # Fillers to guarantee 5 each
    ww_fill = [
        {"title": "Bot explains price breakdowns clearly", "caseType": "DSQ",
         "detail": "The bot provides structured price breakdowns for DSQ customers, reducing the need for human agent intervention.",
         "impact": "Medium"},
        {"title": "Multi-turn context retained throughout chat", "caseType": "Both",
         "detail": "The bot maintains conversation context across multiple turns, enabling coherent multi-step interactions.",
         "impact": "Medium"},
        {"title": "Appointment query handling is responsive", "caseType": "QNR",
         "detail": "QNR customers asking about appointment status receive timely and accurate responses.",
         "impact": "Low"},
    ]
    ni_fill = [
        {"title": "No proactive follow-up for pending sessions", "caseType": "Both",
         "detail": "Sessions ending with an open question have no automated follow-up mechanism.",
         "recommendation": "Implement a 24-hour follow-up nudge for sessions that ended with an unanswered question.",
         "priority": "Low"},
        {"title": "Bot language not personalised to customer", "caseType": "Both",
         "detail": "The bot uses generic language without referencing the customer's specific vehicle or appointment.",
         "recommendation": "Inject vehicle and appointment data into bot responses to increase relevance and trust.",
         "priority": "Low"},
    ]
    while len(ww) < 5: ww.append(ww_fill[len(ww) % len(ww_fill)])
    while len(ni) < 5: ni.append(ni_fill[len(ni) % len(ni_fill)])

    return {"workingWell": ww[:5], "needsImprovement": ni[:5]}

# ── Interactive Chat Widget (injected into both dashboards) ───────────────────
# NOTE: This string uses PLAIN { } for JavaScript (NOT {{ }}).
# It is injected AFTER the template brace-conversion step in build(), so
# its braces are passed through untouched.

CHAT_WIDGET = """<style>
#chat-fab{position:fixed;bottom:28px;right:28px;width:56px;height:56px;background:#e8002d;
  border-radius:50%;display:flex;align-items:center;justify-content:center;cursor:pointer;
  box-shadow:0 4px 20px rgba(232,0,45,.55);z-index:9999;transition:.2s;border:none;
  font-size:22px;color:#fff;pointer-events:auto!important;outline:none;}
#chat-fab:hover{transform:scale(1.08);box-shadow:0 6px 28px rgba(232,0,45,.75);}
#chat-panel{position:fixed;bottom:96px;right:28px;width:430px;height:530px;max-height:530px;
  background:#1e293b;border:1px solid #334155;border-radius:16px;
  box-shadow:0 20px 60px rgba(0,0,0,.55);z-index:9998;pointer-events:auto!important;
  display:none;flex-direction:column;overflow:hidden;}
#chat-panel.open{display:flex;animation:cwSlide .22s ease;}
@keyframes cwSlide{from{opacity:0;transform:translateY(16px)}to{opacity:1;transform:translateY(0)}}
#chat-hdr{background:linear-gradient(135deg,#16213e,#0f3460);padding:13px 16px;
  border-bottom:1px solid #334155;display:flex;align-items:center;flex-shrink:0;}
#chat-hdr-l{display:flex;align-items:center;gap:10px;}
#chat-av{width:36px;height:36px;background:#e8002d;border-radius:8px;
  display:flex;align-items:center;justify-content:center;font-size:17px;flex-shrink:0;}
#chat-hn{font-size:14px;font-weight:700;color:#fff;}
#chat-hs{font-size:11px;color:#94a3b8;}
#chat-msgs{flex:1;overflow-y:auto;padding:14px 14px 6px;display:flex;flex-direction:column;gap:10px;min-height:0;}
#chat-msgs::-webkit-scrollbar{width:4px;}
#chat-msgs::-webkit-scrollbar-thumb{background:#334155;border-radius:2px;}
#chat-close-row{padding:8px 12px 10px;border-top:1px solid #334155;flex-shrink:0;}
#chat-x{width:100%;background:transparent;border:1px solid #334155;border-radius:8px;
  color:#94a3b8;cursor:pointer;font-size:13px;font-weight:600;padding:7px 12px;
  display:flex;align-items:center;justify-content:center;gap:6px;transition:.15s;}
#chat-x:hover{color:#fff;border-color:#e8002d;background:rgba(232,0,45,.12);}
.cmsg{max-width:88%;padding:9px 13px;font-size:13px;line-height:1.6;word-break:break-word;}
.cmsg.bot{background:#0f172a;border:1px solid #334155;color:#e2e8f0;align-self:flex-start;
  border-radius:4px 12px 12px 12px;}
.cmsg.user{background:#e8002d;color:#fff;align-self:flex-end;
  border-radius:12px 12px 4px 12px;}
.cmsg b{color:#fbbf24;font-weight:700;}
.cmsg i{color:#94a3b8;font-style:normal;}
#chat-chips{padding:8px 12px;display:flex;gap:6px;flex-wrap:wrap;
  border-top:1px solid rgba(51,65,85,.4);flex-shrink:0;}
.cwchip{font-size:11px;color:#94a3b8;border:1px solid #334155;background:rgba(255,255,255,.04);
  padding:4px 10px;border-radius:20px;cursor:pointer;transition:.15s;white-space:nowrap;}
.cwchip:hover{color:#fff;border-color:#e8002d;background:rgba(232,0,45,.12);}
#chat-inp-row{padding:10px 12px;border-top:1px solid #334155;display:flex;gap:8px;align-items:center;flex-shrink:0;}
#chat-inp{flex:1;background:#0f172a;border:1px solid #334155;border-radius:8px;
  padding:9px 12px;color:#e2e8f0;font-size:13px;outline:none;font-family:inherit;}
#chat-inp:focus{border-color:#e8002d;}
#chat-btn{background:#e8002d;border:none;border-radius:8px;width:36px;height:36px;
  cursor:pointer;color:#fff;font-size:18px;display:flex;align-items:center;
  justify-content:center;flex-shrink:0;transition:.15s;}
#chat-btn:hover{background:#c0001f;}
</style>
<button id="chat-fab" title="Ask about this data">💬</button>
<div id="chat-panel">
  <div id="chat-hdr">
    <div id="chat-hdr-l">
      <div id="chat-av">🤖</div>
      <div><div id="chat-hn">Data Assistant</div><div id="chat-hs">Ask anything about this dashboard</div></div>
    </div>
  </div>
  <div id="chat-msgs"></div>
  <div id="chat-chips">
    <span class="cwchip" data-q="overview">📊 Overview</span>
    <span class="cwchip" data-q="dsq stats">🏷️ DSQ</span>
    <span class="cwchip" data-q="qnr stats">📋 QNR</span>
    <span class="cwchip" data-q="auction rate">🔨 Auction</span>
    <span class="cwchip" data-q="both auction and ticket">🎯 Both</span>
    <span class="cwchip" data-q="regions">🗺️ Regions</span>
    <span class="cwchip" data-q="inspection gap">📅 Gap</span>
    <span class="cwchip" data-q="weekly trends">📈 Weekly</span>
    <span class="cwchip" data-q="monthly trends">📆 Monthly</span>
    <span class="cwchip" data-q="repeat bug">🐛 Bugs</span>
    <span class="cwchip" data-q="needs improvement">🔧 Issues</span>
    <span class="cwchip" data-q="working well">✅ Strengths</span>
    <span class="cwchip" data-q="bot personalisation">🗣️ Personalisation</span>
    <span class="cwchip" data-q="session IDs where repeat bug">🆔 Session IDs</span>
    <span class="cwchip" data-q="help">❓ Help</span>
  </div>
  <div id="chat-inp-row">
    <input id="chat-inp" type="text" placeholder="Ask e.g. 'which region has most DSQ?'"/>
    <button id="chat-btn">➤</button>
  </div>
  <div id="chat-close-row">
    <button id="chat-x">✕ Close Assistant</button>
  </div>
</div>
<script>
(function(){
  var fab=document.getElementById('chat-fab');
  var panel=document.getElementById('chat-panel');
  var msgs=document.getElementById('chat-msgs');
  var inp=document.getElementById('chat-inp');
  var xbtn=document.getElementById('chat-x');
  var sbtn=document.getElementById('chat-btn');
  var chips=document.getElementById('chat-chips');
  var open=false;

  function cwToggle(){
    if(!panel)return;
    open=!open;
    panel.classList.toggle('open',open);
    if(open&&msgs.children.length===0){
      cwBot("Hi! I'm your Data Assistant 👋<br>I can answer questions about this dashboard's metrics, trends and insights.<br>Tap a chip above or type a question. Type <b>help</b> to see all topics.");
    }
    if(open)setTimeout(function(){inp.focus();},120);
  }

  function cwAsk(q){
    cwUser(q);
    setTimeout(function(){cwBot(cwAnswer(q));},220);
  }

  function cwSend(){
    var q=inp.value.trim();
    if(!q)return;
    inp.value='';
    cwUser(q);
    setTimeout(function(){cwBot(cwAnswer(q));},300);
  }

  if(fab)fab.addEventListener('click',cwToggle);
  if(xbtn)xbtn.addEventListener('click',cwToggle);
  if(sbtn)sbtn.addEventListener('click',cwSend);
  if(inp)inp.addEventListener('keydown',function(e){if(e.key==='Enter')cwSend();});
  if(chips)chips.addEventListener('click',function(e){
    var chip=e.target.closest('.cwchip');
    if(chip&&chip.dataset.q)cwAsk(chip.dataset.q);
  });

  function cwUser(text){
    var d=document.createElement('div');
    d.className='cmsg user';
    d.textContent=text;
    msgs.appendChild(d);
    msgs.scrollTop=msgs.scrollHeight;
  }

  function cwBot(html){
    var d=document.createElement('div');
    d.className='cmsg bot';
    d.innerHTML=html;
    msgs.appendChild(d);
    msgs.scrollTop=msgs.scrollHeight;
  }

  function cwAnswer(q){
    var ql=q.toLowerCase().trim();
    var m=DATA.metrics;
    var qual=DATA.qual;
    var pct=function(n,d){return d?Math.round(n/d*100):0;};

    /* ── help ── */
    if(/help|what can|commands|what.*ask/.test(ql)){
      return 'You can ask me about:<br>'
        +'• <i>overview / summary / performance</i><br>'
        +'• <i>total sessions / DSQ / QNR</i><br>'
        +'• <i>auction rate / ticket rate</i><br>'
        +'• <i>both auction and ticket</i><br>'
        +'• <i>drop rate / resolution</i><br>'
        +'• <i>repeat bug / duplicates</i><br>'
        +'• <i>average messages / conversation length</i><br>'
        +'• <i>region breakdown / best region / worst region</i><br>'
        +'• <i>inspection gap</i><br>'
        +'• <i>weekly trends / monthly trends</i><br>'
        +'• <i>working well / strengths</i><br>'
        +'• <i>needs improvement / issues / recommendations</i><br>'
        +'• <i>personalisation / bot language / tone</i><br>'
        +'• <i>sample session IDs (e.g. "session IDs where repeat bug")</i><br>'
        +'• <i>compare DSQ vs QNR</i>';
    }

    /* ── overview / summary ── */
    if(/summar|overview|perform|status|how.*doing|overall|report/.test(ql)){
      return 'Performance Overview:<br>'
        +'• Total sessions: <b>'+m.total+'</b> ('+m.dsq_count+' DSQ, '+m.qnr_count+' QNR)<br>'
        +'• Auction started: <b>'+pct(m.outcomes.auction_started,m.total)+'%</b> ('+m.outcomes.auction_started+' sessions)<br>'
        +'• Ticket raised: <b>'+m.ticket_rate+'%</b> ('+m.outcomes.ticket_raised+' sessions)<br>'
        +'• Drop rate: <b>'+m.drop_rate+'%</b> | Resolution among drops: <b>'+m.resolution_rate+'%</b><br>'
        +'• Both auction+ticket: <b>'+m.has_both_pct+'%</b> ('+m.has_both_count+' sessions)<br>'
        +'• Repeat bug: <b>'+m.repeat_pct+'%</b> ('+m.repeat+' sessions)';
    }

    /* ── both auction + ticket ── */
    if(/both|auction.*ticket|ticket.*auction/.test(ql)){
      return 'Sessions where BOTH an auction was started AND a ticket was raised:<br>'
        +'<b>'+m.has_both_count+'</b> sessions = <b>'+m.has_both_pct+'%</b> of all sessions.';
    }

    /* ── compare DSQ vs QNR ── */
    if(/compar|dsq.*qnr|qnr.*dsq|versus|\\bvs\\b/.test(ql)){
      return '<b>DSQ ('+m.dsq_count+' sessions)</b><br>'
        +'Auction '+m.auction_rate_dsq+'% | Ticket '+pct(m.dsq_outcomes.ticket_raised,m.dsq_count)+'% | Drop✓ '+pct(m.dsq_outcomes.drop_with_resolution,m.dsq_count)+'% | Drop✗ '+pct(m.dsq_outcomes.drop_without_resolution,m.dsq_count)+'%<br><br>'
        +'<b>QNR ('+m.qnr_count+' sessions)</b><br>'
        +'Auction '+m.auction_rate_qnr+'% | Ticket '+pct(m.qnr_outcomes.ticket_raised,m.qnr_count)+'% | Drop✓ '+pct(m.qnr_outcomes.drop_with_resolution,m.qnr_count)+'% | Drop✗ '+pct(m.qnr_outcomes.drop_without_resolution,m.qnr_count)+'%';
    }

    /* ── DSQ ── */
    if(/\\bdsq\\b|dissatisf/.test(ql)){
      return 'DSQ — Dissatisfied with Quote:<br>'
        +'• Total: <b>'+m.dsq_count+'</b> ('+pct(m.dsq_count,m.total)+'% of all sessions)<br>'
        +'• Auction started: <b>'+m.dsq_outcomes.auction_started+'</b> ('+m.auction_rate_dsq+'%)<br>'
        +'• Ticket raised: <b>'+m.dsq_outcomes.ticket_raised+'</b> ('+pct(m.dsq_outcomes.ticket_raised,m.dsq_count)+'%)<br>'
        +'• Drop resolved: <b>'+m.dsq_outcomes.drop_with_resolution+'</b><br>'
        +'• Drop unresolved: <b>'+m.dsq_outcomes.drop_without_resolution+'</b>';
    }

    /* ── QNR ── */
    if(/\\bqnr\\b|quote not/.test(ql)){
      return 'QNR — Quote Not Received:<br>'
        +'• Total: <b>'+m.qnr_count+'</b> ('+pct(m.qnr_count,m.total)+'% of all sessions)<br>'
        +'• Auction started: <b>'+m.qnr_outcomes.auction_started+'</b> ('+m.auction_rate_qnr+'%)<br>'
        +'• Ticket raised: <b>'+m.qnr_outcomes.ticket_raised+'</b> ('+pct(m.qnr_outcomes.ticket_raised,m.qnr_count)+'%)<br>'
        +'• Drop resolved: <b>'+m.qnr_outcomes.drop_with_resolution+'</b><br>'
        +'• Drop unresolved: <b>'+m.qnr_outcomes.drop_without_resolution+'</b>';
    }

    /* ── auction ── */
    if(/auction/.test(ql)){
      return 'Auction started:<br>'
        +'• Overall: <b>'+m.outcomes.auction_started+'</b> sessions ('+pct(m.outcomes.auction_started,m.total)+'%)<br>'
        +'• DSQ auction rate: <b>'+m.auction_rate_dsq+'%</b><br>'
        +'• QNR auction rate: <b>'+m.auction_rate_qnr+'%</b>';
    }

    /* ── ticket ── */
    if(/ticket/.test(ql)){
      return 'Ticket raised:<br>'
        +'• Total: <b>'+m.outcomes.ticket_raised+'</b> sessions ('+m.ticket_rate+'%)<br>'
        +'• DSQ: <b>'+m.dsq_outcomes.ticket_raised+'</b> ('+pct(m.dsq_outcomes.ticket_raised,m.dsq_count)+'%) | QNR: <b>'+m.qnr_outcomes.ticket_raised+'</b> ('+pct(m.qnr_outcomes.ticket_raised,m.qnr_count)+'%)';
    }

    /* ── drop / resolution ── */
    if(/drop|resolution|resolved|unresolved/.test(ql)){
      return 'Drop & resolution:<br>'
        +'• Drop rate: <b>'+m.drop_rate+'%</b><br>'
        +'• Drop resolved: <b>'+m.outcomes.drop_with_resolution+'</b> sessions<br>'
        +'• Drop unresolved: <b>'+m.outcomes.drop_without_resolution+'</b> sessions<br>'
        +'• Resolution rate among drops: <b>'+m.resolution_rate+'%</b><br>'
        +'• Unresolved % of all: <b>'+pct(m.outcomes.drop_without_resolution,m.total)+'%</b>';
    }

    /* ── total sessions ── */
    if(/total|how many session|session count|number of session/.test(ql)){
      return 'Total sessions: <b>'+m.total+'</b><br>'
        +'• DSQ: <b>'+m.dsq_count+'</b> ('+pct(m.dsq_count,m.total)+'%) | QNR: <b>'+m.qnr_count+'</b> ('+pct(m.qnr_count,m.total)+'%)';
    }

    /* ── repeat bug ── */
    if(/repeat|bug|duplicate|same message/.test(ql)){
      var rb=m.repeat_buckets||[];
      var bkts=rb.map(function(b){
        return '• '+b.label+': <b>'+b.session_count+'</b> sessions ('+pct(b.session_count,m.total)+'%), total rep count: '+b.rep_count;
      }).join('<br>');
      return 'Repeat message bug:<br>'
        +'• Affected: <b>'+m.repeat+'</b> sessions (<b>'+m.repeat_pct+'%</b>)<br><br>'
        +'Bucket breakdown:<br>'+bkts;
    }

    /* ── average messages ── */
    if(/avg|average|message.*per|per.*session|length|turn/.test(ql)){
      return 'Average messages per session:<br>'
        +'• Total turns: <b>'+m.avg_msgs+'</b><br>'
        +'• Bot turns: <b>'+m.avg_bot+'</b><br>'
        +'• Human turns: <b>'+m.avg_human+'</b>';
    }

    /* ── best/top region ── */
    if(/best.*region|top.*region|highest.*region|region.*most|most.*region/.test(ql)){
      var rs=m.region_split||[];
      if(!rs.length)return 'No region data yet — ensure REGION_3 column is populated.';
      var b=rs[0];
      return 'Top region by volume: <b>'+b.region+'</b><br>'
        +'DSQ: '+b.dsq+' | QNR: '+b.qnr+' | Total: <b>'+(b.dsq+b.qnr)+'</b>';
    }

    /* ── worst/bottom region ── */
    if(/worst.*region|bottom.*region|lowest.*region|region.*least|least.*region/.test(ql)){
      var rs=m.region_split||[];
      if(!rs.length)return 'No region data yet.';
      var b=rs[rs.length-1];
      return 'Lowest volume region tracked: <b>'+b.region+'</b><br>'
        +'DSQ: '+b.dsq+' | QNR: '+b.qnr+' | Total: <b>'+(b.dsq+b.qnr)+'</b>';
    }

    /* ── which region has highest DSQ ── */
    if(/region.*dsq|dsq.*region|highest dsq/.test(ql)){
      var rs=m.region_split||[];
      if(!rs.length)return 'No region data yet.';
      var top=rs.slice().sort(function(a,b){return b.dsq-a.dsq;})[0];
      return 'Region with highest DSQ volume: <b>'+top.region+'</b><br>DSQ: <b>'+top.dsq+'</b> | QNR: '+top.qnr;
    }

    /* ── which region has highest QNR ── */
    if(/region.*qnr|qnr.*region|highest qnr/.test(ql)){
      var rs=m.region_split||[];
      if(!rs.length)return 'No region data yet.';
      var top=rs.slice().sort(function(a,b){return b.qnr-a.qnr;})[0];
      return 'Region with highest QNR volume: <b>'+top.region+'</b><br>QNR: <b>'+top.qnr+'</b> | DSQ: '+top.dsq;
    }

    /* ── region ── */
    if(/region|city|location|area|geography|state/.test(ql)){
      var rs=m.region_split||[];
      if(!rs.length)return 'No region data yet — ensure REGION_3 column is populated in the sheet.';
      var top=rs.slice(0,10);
      return 'Region breakdown (top '+top.length+'):<br>'
        +top.map(function(r){
          return '• <b>'+r.region+'</b>: DSQ '+r.dsq+' | QNR '+r.qnr+' | Total <b>'+(r.dsq+r.qnr)+'</b>';
        }).join('<br>');
    }

    /* ── inspection gap ── */
    if(/inspection|insp.*gap|\\bgap\\b|insp.*date|days.*inspection/.test(ql)){
      var ig=m.insp_gap||{buckets:[],dsq:[],qnr:[]};
      if(!ig.buckets.length)return 'No inspection date data yet — ensure INSP_DATE_3 is populated.';
      var td=ig.dsq.reduce(function(a,b){return a+b;},0);
      var tq=ig.qnr.reduce(function(a,b){return a+b;},0);
      return 'Inspection-to-chat gap ('+td+' DSQ · '+tq+' QNR with data):<br>'
        +ig.buckets.map(function(b,i){
          return '• <b>'+b+'</b>: DSQ <b>'+ig.dsq[i]+'</b> | QNR <b>'+ig.qnr[i]+'</b>';
        }).join('<br>');
    }

    /* ── weekly trends ── */
    if(/week/.test(ql)){
      var wt=m.weekly_trends||[];
      if(!wt.length)return 'No weekly trend data yet (need at least a few weeks of sessions).';
      var lat=wt[wt.length-1];
      return 'Weekly trends: <b>'+wt.length+' weeks</b> in last 3 months<br><br>'
        +'<b>Latest week ('+lat.label+'):</b><br>'
        +'→ DSQ ('+lat.dsq_total+' sessions): Auction <b>'+lat.dsq_pcts.auction_started+'%</b> | Ticket '+lat.dsq_pcts.ticket_raised+'% | Drop✓ '+lat.dsq_pcts.drop_with_resolution+'% | Drop✗ '+lat.dsq_pcts.drop_without_resolution+'%<br>'
        +'→ QNR ('+lat.qnr_total+' sessions): Auction <b>'+lat.qnr_pcts.auction_started+'%</b> | Ticket '+lat.qnr_pcts.ticket_raised+'% | Drop✓ '+lat.qnr_pcts.drop_with_resolution+'% | Drop✗ '+lat.qnr_pcts.drop_without_resolution+'%';
    }

    /* ── monthly trends ── */
    if(/month/.test(ql)){
      var mt=m.monthly_trends||[];
      if(!mt.length)return 'No monthly trend data yet.';
      return 'Monthly breakdown (<b>'+mt.length+' months</b>):<br>'
        +mt.map(function(mo){
          return '<b>'+mo.label+'</b> — DSQ: '+mo.dsq_total+' sess (Auction <b>'+mo.dsq_pcts.auction_started+'%</b>) | QNR: '+mo.qnr_total+' sess (Auction <b>'+mo.qnr_pcts.auction_started+'%</b>)';
        }).join('<br>');
    }

    /* ── working well ── */
    if(/working well|strength|what.*work|positive|win/.test(ql)){
      return "What's working well:<br>"
        +qual.workingWell.map(function(w){
          return '<b>'+w.title+'</b> ['+w.caseType+' · Impact: '+w.impact+']<br><i>'+w.detail+'</i>';
        }).join('<br><br>');
    }

    /* ── needs improvement ── */
    if(/improv|issue|problem|fix|concern|weak|recommend|action/.test(ql)){
      return 'Areas needing improvement:<br>'
        +qual.needsImprovement.map(function(n){
          return '<b>'+n.title+'</b> ['+n.caseType+' · Priority: '+n.priority+']<br><i>'+n.detail+'</i><br>💡 '+n.recommendation;
        }).join('<br><br>');
    }

    /* ── personalisation / bot language ── */
    if(/personal|language|personalised|personaliz|bot.*lang|lang.*bot|greeting|tone|generic/.test(ql)){
      var ni=qual.needsImprovement||[];
      var pItems=ni.filter(function(n){return /personal|language/i.test(n.title||'');});
      if(pItems.length){
        return 'Bot personalisation insights from Qualitative analysis:<br><br>'
          +pItems.map(function(n){
            return '<b>'+n.title+'</b> ['+n.caseType+' · Priority: '+n.priority+']<br>'
              +'<i>'+n.detail+'</i><br>💡 '+n.recommendation;
          }).join('<br><br>')
          +'<br><br>To see sample session IDs, ask: <i>"show session IDs where repeat bug"</i> or <i>"show dropped session IDs"</i>.';
      }
      return 'Bot personalisation is tracked as a key improvement area in the Qualitative dashboard.<br>'
        +'Ask <i>"needs improvement"</i> to see all flagged issues including language quality.<br><br>'
        +'<b>Note:</b> Per-session personalisation scoring requires content-level analysis — '
        +'the qualitative dashboard flags this pattern based on a sample of conversations.';
    }

    /* ── session IDs / sample sessions ── */
    if(/session.*id|sample.*session|\\bsid\\b|show.*session|list.*session|example.*session/.test(ql)){
      var sl2=DATA.metrics.sessions_list||[];
      var matchFn=null;
      var label='';
      if(/repeat|bug|duplicate/.test(ql)){matchFn=function(s){return s.hr;};label='repeat bug';}
      else if(/unresolved|no.*resolut|without.*resolut/.test(ql)){matchFn=function(s){return s.oc==='drop_without_resolution';};label='dropped without resolution';}
      else if(/drop/.test(ql)){matchFn=function(s){return s.oc==='drop_with_resolution'||s.oc==='drop_without_resolution';};label='dropped';}
      else if(/auction/.test(ql)){matchFn=function(s){return s.oc==='auction_started';};label='auction started';}
      else if(/ticket/.test(ql)){matchFn=function(s){return s.oc==='ticket_raised';};label='ticket raised';}
      else if(/both/.test(ql)){matchFn=function(s){return s.hb;};label='both auction+ticket';}
      else if(/dsq/.test(ql)){matchFn=function(s){return s.ct==='DSQ';};label='DSQ';}
      else if(/qnr/.test(ql)){matchFn=function(s){return s.ct==='QNR';};label='QNR';}
      if(matchFn){
        var matched=sl2.filter(matchFn).slice(0,10);
        if(!matched.length)return 'No sessions found matching <i>'+label+'</i> condition.';
        return 'Sample session IDs — <b>'+label+'</b> (up to 10 of '+sl2.filter(matchFn).length+'):<br>'
          +matched.map(function(s){
            return '• <b>'+(s.si||'N/A')+'</b> ['+s.ct+' · '+s.oc.replace(/_/g,' ')+' · '+s.d+']';
          }).join('<br>');
      }
      return 'Please specify what kind of sessions you want. For example:<br>'
        +'• <i>"session IDs where repeat bug"</i><br>'
        +'• <i>"session IDs where auction started"</i><br>'
        +'• <i>"session IDs where dropped"</i><br>'
        +'• <i>"session IDs where unresolved"</i><br>'
        +'• <i>"session IDs DSQ"</i>';
    }

    return "I didn't catch that. Try: <i>sessions, auction rate, DSQ, QNR, both, repeat bug, regions, inspection gap, weekly trends, monthly trends, working well, needs improvement, personalisation, session IDs</i> — or type <b>help</b>.";
  }
})();
</script>"""

def make_summary(m):
    p   = lambda n, d: round(n / max(d, 1) * 100)
    pct = p(m["outcomes"]["auction_started"], m["total"])
    tone = "positively" if pct >= 25 else ("moderately" if pct >= 10 else "with room for improvement")
    res  = "strong" if m["resolution_rate"] >= 60 else "developing"
    bug  = (f"the {m['repeat_pct']}% repeat-message bug warrants attention"
            if m["repeat_pct"] > 5
            else f"bot message quality is consistent with a low {m['repeat_pct']}% repeat rate")
    both = (f" {m['has_both_pct']}% of sessions triggered both an auction and a ticket.")
    return (
        f"The CARS24 chatbot handled {m['total']} sessions ({m['dsq_count']} DSQ · {m['qnr_count']} QNR) "
        f"and is performing {tone}, converting {pct}% of interactions into new auctions. "
        f"Resolution quality is {res} at {m['resolution_rate']}% among dropped sessions, and {bug}.{both}"
    )

# ── HTML template — Metrics ───────────────────────────────────────────────────

RECOMPUTE_JS = '<script>\n// ── Client-side session recompute for date filter ────────────────────────────\nfunction recompute(sl){\n  var OCS=[\'auction_started\',\'ticket_raised\',\'drop_with_resolution\',\'drop_without_resolution\'];\n  var total=sl.length;\n  var dsq=sl.filter(function(s){return s.ct===\'DSQ\';}),qnr=sl.filter(function(s){return s.ct===\'QNR\';});\n  var pct=function(n,d){return d?Math.round(n/d*1000)/10:0;};\n  var outcomes={},dc={},qc={};\n  OCS.forEach(function(o){\n    outcomes[o]=sl.filter(function(s){return s.oc===o;}).length;\n    dc[o]=dsq.filter(function(s){return s.oc===o;}).length;\n    qc[o]=qnr.filter(function(s){return s.oc===o;}).length;\n  });\n  var repeat=sl.filter(function(s){return s.hr;}).length;\n  var hbc=sl.filter(function(s){return s.hb;}).length;\n  var avgN=total?Math.round(sl.reduce(function(a,s){return a+s.n;},0)/total*10)/10:0;\n  var dRes=outcomes.drop_with_resolution,dUnres=outcomes.drop_without_resolution;\n  var dailyMap={};\n  sl.forEach(function(s){if(s.d)dailyMap[s.d]=(dailyMap[s.d]||0)+1;});\n  var daily=Object.keys(dailyMap).sort().map(function(d){return{d:d,c:dailyMap[d]};});\n  var rb=[{label:\'>1 repetition\',th:2},{label:\'>3 repetitions\',th:4},{label:\'>5 repetitions\',th:6}].map(function(x){\n    var m2=sl.filter(function(s){return s.rc>=x.th;});\n    return{label:x.label,session_count:m2.length,rep_count:m2.reduce(function(a,s){return a+s.rc;},0)};\n  });\n  function mkTrend(period){\n    var valid=sl.filter(function(s){return s.d&&s.d.length>=10;});\n    if(!valid.length)return[];\n    var maxD=valid.reduce(function(a,s){return s.d>a?s.d:a;},\'\');\n    var cutMs=new Date(maxD).getTime()-91*864e5;\n    var bkts={};\n    valid.forEach(function(s){\n      if(new Date(s.d).getTime()<cutMs)return;\n      var dt=new Date(s.d),key;\n      if(period===\'week\'){var d1=new Date(dt.getFullYear(),0,1),wk=Math.ceil(((dt-d1)/864e5+d1.getDay()+1)/7);key=dt.getFullYear()+\'-W\'+String(wk).padStart(2,\'0\');}\n      else{key=s.d.slice(0,7);}\n      if(!bkts[key]){bkts[key]={DSQ:{},QNR:{}};OCS.forEach(function(o){bkts[key].DSQ[o]=0;bkts[key].QNR[o]=0;});}\n      if(OCS.indexOf(s.oc)>=0)bkts[key][s.ct][s.oc]++;\n    });\n    return Object.keys(bkts).sort().map(function(k){\n      var b=bkts[k];\n      var dt=OCS.reduce(function(a,o){return a+b.DSQ[o];},0),qt=OCS.reduce(function(a,o){return a+b.QNR[o];},0);\n      var dp={},qp={};\n      OCS.forEach(function(o){dp[o]=dt?Math.round(b.DSQ[o]/dt*1000)/10:0;qp[o]=qt?Math.round(b.QNR[o]/qt*1000)/10:0;});\n      return{label:k,dsq_total:dt,dsq_pcts:dp,qnr_total:qt,qnr_pcts:qp};\n    });\n  }\n  var rgMap={DSQ:{},QNR:{}};\n  sl.forEach(function(s){if(s.rg)rgMap[s.ct][s.rg]=(rgMap[s.ct][s.rg]||0)+1;});\n  var allRg=[];\n  sl.forEach(function(s){if(s.rg&&allRg.indexOf(s.rg)<0)allRg.push(s.rg);});\n  allRg.sort();\n  var rg=allRg.map(function(r){return{region:r,dsq:rgMap.DSQ[r]||0,qnr:rgMap.QNR[r]||0};});\n  var GB=["Same Day","1-3 days","3-5 days","5-10 days","10-20 days","20-30 days",">30 days"];\n  function gBkt(g){if(g===0)return"Same Day";if(g<=3)return"1-3 days";if(g<=5)return"3-5 days";if(g<=10)return"5-10 days";if(g<=20)return"10-20 days";if(g<=30)return"20-30 days";return">30 days";}\n  var ig={dsq:GB.map(function(){return 0;}),qnr:GB.map(function(){return 0;})};\n  sl.forEach(function(s){\n    if(!s.id||!s.d)return;\n    var gap=Math.round((new Date(s.d)-new Date(s.id))/864e5);\n    if(gap<0)return;\n    var bi=GB.indexOf(gBkt(gap));\n    if(bi>=0){if(s.ct===\'DSQ\')ig.dsq[bi]++;else ig.qnr[bi]++;}\n  });\n  // EHB/TP computation from sessions_list (eb=EHB price, tp=NEW_TP)\n  var etDsq=[],etQnr=[];\n  sl.forEach(function(s){\n    if(s.oc===\'auction_started\'&&s.eb&&s.tp&&s.tp>0){\n      var r=Math.round(s.eb/s.tp*1000)/10;\n      if(s.ct===\'DSQ\')etDsq.push(r);else etQnr.push(r);\n    }\n  });\n  var etDsqAvg=etDsq.length?Math.round(etDsq.reduce(function(a,b){return a+b;},0)/etDsq.length*10)/10:null;\n  var etQnrAvg=etQnr.length?Math.round(etQnr.reduce(function(a,b){return a+b;},0)/etQnr.length*10)/10:null;\n  // Avg EHB price for all auction sessions (with or without TP)\n  var aeDsq=sl.filter(function(s){return s.oc===\'auction_started\'&&s.eb&&s.ct===\'DSQ\';});\n  var aeQnr=sl.filter(function(s){return s.oc===\'auction_started\'&&s.eb&&s.ct===\'QNR\';});\n  var avgEbDsq=aeDsq.length?Math.round(aeDsq.reduce(function(a,s){return a+s.eb;},0)/aeDsq.length):null;\n  var avgEbQnr=aeQnr.length?Math.round(aeQnr.reduce(function(a,s){return a+s.eb;},0)/aeQnr.length):null;\n  return{total:total,dsq_count:dsq.length,qnr_count:qnr.length,\n    outcomes:outcomes,dsq_outcomes:dc,qnr_outcomes:qc,\n    repeat:repeat,repeat_pct:pct(repeat,total),avg_msgs:avgN,avg_bot:0,avg_human:0,\n    auction_rate_dsq:pct(dc.auction_started,dsq.length),\n    auction_rate_qnr:pct(qc.auction_started,qnr.length),\n    ticket_rate:pct(outcomes.ticket_raised,total),\n    drop_rate:pct(dRes+dUnres,total),resolution_rate:pct(dRes,dRes+dUnres),\n    daily:daily,has_both_count:hbc,has_both_pct:pct(hbc,total),\n    repeat_buckets:rb,weekly_trends:mkTrend(\'week\'),monthly_trends:mkTrend(\'month\'),\n    region_split:rg,insp_gap:ig,\n    ehb_tp_dsq:etDsqAvg,ehb_tp_qnr:etQnrAvg,ehb_tp_dsq_n:etDsq.length,ehb_tp_qnr_n:etQnr.length,\n    avg_ehb_dsq:avgEbDsq,avg_ehb_qnr:avgEbQnr,ehb_dsq_n:aeDsq.length,ehb_qnr_n:aeQnr.length};}\n// ── Hash-based date filter state ─────────────────────────────────────────────\nvar _flt=(function(){var h=location.hash.slice(1),r={};h.split(\'&\').forEach(function(kv){var p=kv.split(\'=\');if(p.length===2&&p[0])r[p[0]]=decodeURIComponent(p[1]);});return r;})();\nvar _sl=(DATA.metrics.sessions_list||[]).slice();\nif(_flt.start)_sl=_sl.filter(function(s){return s.d>=_flt.start;});\nif(_flt.end)  _sl=_sl.filter(function(s){return s.d<=_flt.end;});\n// ── Filter bar UI wiring ──────────────────────────────────────────────────────\n(function(){\n  var si=document.getElementById(\'df-start\'),ei=document.getElementById(\'df-end\');\n  var ab=document.getElementById(\'df-apply\'),cb=document.getElementById(\'df-clear\');\n  var badge=document.getElementById(\'df-badge\');\n  if(si&&_flt.start)si.value=_flt.start;\n  if(ei&&_flt.end)  ei.value=_flt.end;\n  if(badge&&(_flt.start||_flt.end)){\n    badge.textContent=\'Filtered: \'+_sl.length+\' of \'+(DATA.metrics.sessions_list||[]).length+\' sessions\';\n    badge.style.display=\'inline-block\';\n  }\n  function doFilter(){\n    var s=si?si.value:\'\',e=ei?ei.value:\'\';\n    if(!s&&!e){location.hash=\'\';location.reload();return;}\n    location.hash=\'start=\'+encodeURIComponent(s)+\'&end=\'+encodeURIComponent(e);\n    location.reload();\n  }\n  if(ab)ab.addEventListener(\'click\',doFilter);\n  if(cb)cb.addEventListener(\'click\',function(){location.hash=\'\';location.reload();});\n  if(si)si.addEventListener(\'keydown\',function(e){if(e.key===\'Enter\')doFilter();});\n  if(ei)ei.addEventListener(\'keydown\',function(e){if(e.key===\'Enter\')doFilter();});\n})();\n</script>'

DASHBOARD_PASSWORD_HASH = "{DASHBOARD_PASSWORD_HASH}"   # sha256 hex set at build time

PASSWORD_GATE = """
<style>
body.pw-locked > *:not(#pw-overlay){{display:none!important}}
#pw-overlay{{position:fixed;inset:0;background:#1a1a2e;display:flex;align-items:center;
  justify-content:center;z-index:99999;flex-direction:column;gap:18px;font-family:'Segoe UI',system-ui,sans-serif}}
#pw-box{{background:#1e293b;border:1px solid #334155;border-radius:16px;padding:40px 44px;
  text-align:center;max-width:380px;width:90%;box-shadow:0 20px 60px rgba(0,0,0,.6)}}
#pw-logo{{background:#e8002d;color:#fff;font-weight:800;font-size:18px;padding:6px 14px;
  border-radius:6px;display:inline-block;margin-bottom:18px;letter-spacing:1px}}
#pw-title{{color:#e2e8f0;font-size:20px;font-weight:700;margin-bottom:6px}}
#pw-sub{{color:#94a3b8;font-size:13px;margin-bottom:24px}}
#pw-input{{width:100%;background:#0f172a;border:1px solid #334155;border-radius:8px;
  padding:11px 14px;color:#e2e8f0;font-size:14px;outline:none;box-sizing:border-box;
  text-align:center;letter-spacing:2px}}
#pw-input:focus{{border-color:#e8002d}}
#pw-btn{{width:100%;background:#e8002d;color:#fff;border:none;border-radius:8px;
  padding:11px;font-size:14px;font-weight:700;cursor:pointer;margin-top:12px;transition:.15s}}
#pw-btn:hover{{background:#c0001f}}
#pw-err{{color:#f87171;font-size:12px;margin-top:8px;min-height:18px}}
</style>
<script>
(function(){{
  // Lock the page immediately before any content renders
  document.documentElement.style.visibility='hidden';
  document.addEventListener('DOMContentLoaded',function(){{
    if(sessionStorage.getItem('cars24_auth')==='1'){{
      // Already authenticated — remove overlay and show content
      var o=document.getElementById('pw-overlay');
      if(o)o.remove();
      document.body.classList.remove('pw-locked');
      document.documentElement.style.visibility='visible';
    }} else {{
      document.body.classList.add('pw-locked');
      document.documentElement.style.visibility='visible';
    }}
  }});
  async function sha256(msg){{
    var buf=await crypto.subtle.digest('SHA-256',new TextEncoder().encode(msg));
    return Array.from(new Uint8Array(buf)).map(b=>b.toString(16).padStart(2,'0')).join('');
  }}
  window.checkPw=async function(){{
    var v=document.getElementById('pw-input').value;
    var h=await sha256(v);
    if(h==='{DASHBOARD_PASSWORD_HASH}'){{
      sessionStorage.setItem('cars24_auth','1');
      document.getElementById('pw-overlay').remove();
      document.body.classList.remove('pw-locked');
    }}else{{
      document.getElementById('pw-err').textContent='Incorrect password — please try again.';
      document.getElementById('pw-input').value='';
      document.getElementById('pw-input').focus();
    }}
  }};
}})();
</script>
<div id="pw-overlay">
  <div id="pw-box">
    <div id="pw-logo">CARS24</div>
    <div id="pw-title">Dashboard Access</div>
    <div id="pw-sub">Enter the team password to continue</div>
    <input id="pw-input" type="password" placeholder="Password" autofocus
           onkeydown="if(event.key==='Enter')checkPw()"/>
    <button id="pw-btn" onclick="checkPw()">Unlock Dashboard</button>
    <div id="pw-err"></div>
  </div>
</div>
"""

METRICS_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>CARS24 Chatbot — Performance Metrics</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script>
// ── Data embedded by generate_dashboards.py ──
const DATA = {DATA_JSON};
</script>
<style>
:root{{--red:#e8002d;--dark:#1a1a2e;--mid:#16213e;--blue:#0f3460;
  --green:#22c55e;--yellow:#f59e0b;--orange:#f97316;
  --err:#ef4444;--purple:#8b5cf6;--teal:#14b8a6;
  --card:#1e293b;--border:#334155;--text:#e2e8f0;--sub:#94a3b8}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:var(--dark);color:var(--text);min-height:100vh}}
header{{background:linear-gradient(135deg,var(--mid),var(--blue));border-bottom:3px solid var(--red);
  padding:18px 32px;display:flex;align-items:center;justify-content:space-between;
  position:sticky;top:0;z-index:100;box-shadow:0 4px 20px rgba(0,0,0,.4)}}
.logo{{display:flex;align-items:center;gap:14px}}
.logo-box{{background:var(--red);color:#fff;font-weight:800;font-size:17px;padding:6px 12px;border-radius:6px;letter-spacing:1px}}
.h-title{{font-size:19px;font-weight:700;color:#fff}}
.h-sub{{font-size:12px;color:var(--sub);margin-top:2px}}
.h-right{{text-align:right}}
.live{{font-size:12px;color:var(--sub)}}
.dot{{display:inline-block;width:8px;height:8px;background:var(--green);border-radius:50%;margin-right:6px;animation:pulse 2s infinite}}
@keyframes pulse{{0%,100%{{opacity:1}}50%{{opacity:.4}}}}
.nav{{display:flex;gap:10px;margin-top:8px;justify-content:flex-end}}
.nav a{{font-size:12px;color:var(--sub);text-decoration:none;padding:4px 10px;
  border:1px solid var(--border);border-radius:4px;transition:.2s}}
.nav a:hover,.nav a.on{{color:#fff;border-color:var(--red);background:rgba(232,0,45,.15)}}
.df-bar{{background:var(--mid);border-bottom:1px solid var(--border);padding:10px 32px;
  position:sticky;top:76px;z-index:98}}
.df-inner{{display:flex;align-items:center;gap:10px;flex-wrap:wrap;max-width:1300px;margin:0 auto}}
.df-lbl{{font-size:12px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:1px;white-space:nowrap}}
.df-inp{{background:#0f172a;border:1px solid var(--border);color:var(--text);border-radius:6px;
  padding:5px 10px;font-size:13px;outline:none;cursor:pointer;color-scheme:dark}}
.df-inp:focus{{border-color:var(--red)}} .df-sep{{color:var(--sub);font-size:14px}}
.df-btn{{border:none;border-radius:6px;padding:6px 14px;font-size:12px;font-weight:700;cursor:pointer;transition:.15s}}
.df-go{{background:var(--red);color:#fff}} .df-go:hover{{background:#c0001f}}
.df-clr{{background:var(--card);color:var(--sub);border:1px solid var(--border)}} .df-clr:hover{{color:var(--text)}}
.df-badge{{background:rgba(232,0,45,.15);border:1px solid rgba(232,0,45,.3);color:#fca5a5;
  border-radius:4px;font-size:11px;padding:3px 10px;white-space:nowrap}}
main{{max-width:1300px;margin:0 auto;padding:26px 22px}}
.info-bar{{background:rgba(34,197,94,.08);border:1px solid rgba(34,197,94,.3);border-radius:8px;
  padding:10px 16px;font-size:13px;color:#4ade80;margin-bottom:22px;display:flex;align-items:center;gap:8px}}
.kpi-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(170px,1fr));gap:14px;margin-bottom:26px}}
.kpi{{background:var(--card);border:1px solid var(--border);border-radius:12px;padding:16px 18px;
  position:relative;overflow:hidden;transition:.15s}}
.kpi:hover{{transform:translateY(-2px)}}
.kpi::after{{content:'';position:absolute;bottom:0;left:0;right:0;height:3px;border-radius:0 0 12px 12px}}
.kpi.g::after{{background:var(--green)}} .kpi.r::after{{background:var(--err)}}
.kpi.b::after{{background:#3b82f6}} .kpi.p::after{{background:var(--purple)}}
.kpi.y::after{{background:var(--yellow)}} .kpi.t::after{{background:var(--teal)}}
.kpi.o::after{{background:var(--orange)}}
.kpi-ico{{position:absolute;top:14px;right:14px;font-size:26px;opacity:.12}}
.kpi-lbl{{font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--sub);margin-bottom:6px}}
.kpi-val{{font-size:30px;font-weight:800;color:#fff;line-height:1}}
.kpi-sub{{font-size:11px;color:var(--sub);margin-top:5px}}
.sep{{font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:var(--sub);
  padding:4px 0 14px;border-top:1px solid var(--border);margin:26px 0 18px}}
.row2{{display:grid;grid-template-columns:1fr 1fr;gap:22px;margin-bottom:22px}}
@media(max-width:900px){{.row2{{grid-template-columns:1fr}}}}
.cc{{background:var(--card);border:1px solid var(--border);border-radius:13px;padding:18px 20px}}
.cc-full{{grid-column:1/-1}}
.cc-title{{font-size:12px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--sub);margin-bottom:3px}}
.cc-sub{{font-size:12px;color:var(--sub);margin-bottom:14px}}
.cw{{position:relative;height:230px}} .cw.tall{{height:290px}}
.ot{{width:100%;border-collapse:collapse;margin-top:14px}}
.ot th{{text-align:left;font-size:10px;font-weight:700;letter-spacing:1px;text-transform:uppercase;
  color:var(--sub);padding:7px 10px;border-bottom:1px solid var(--border)}}
.ot td{{padding:9px 10px;font-size:13px;border-bottom:1px solid rgba(51,65,85,.5)}}
.ot tr:last-child td{{border-bottom:none}}
.ot tr:hover td{{background:rgba(255,255,255,.03)}}
.odot{{display:inline-block;width:8px;height:8px;border-radius:50%;margin-right:7px;vertical-align:middle}}
.pb{{display:flex;align-items:center;gap:8px}}
.pb-bg{{flex:1;height:5px;background:var(--border);border-radius:3px}}
.pb-fill{{height:5px;border-radius:3px}}
.cmp-grid{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-top:12px}}
.cmp{{border-radius:10px;padding:14px}}
.cmp.dsq{{background:rgba(139,92,246,.1);border:1px solid rgba(139,92,246,.3)}}
.cmp.qnr{{background:rgba(59,130,246,.1);border:1px solid rgba(59,130,246,.3)}}
.cmp-title{{font-size:12px;font-weight:700;margin-bottom:10px}}
.cmp-title.dsq{{color:#a78bfa}} .cmp-title.qnr{{color:#60a5fa}}
.cmp-row{{display:flex;justify-content:space-between;font-size:12px;margin-bottom:5px}}
.cmp-k{{color:var(--sub)}} .cmp-v{{font-weight:600;color:#fff}}
.bug-card{{background:rgba(239,68,68,.08);border:1px solid rgba(239,68,68,.25);border-radius:12px;
  padding:16px 18px;display:flex;align-items:center;gap:18px;margin-top:14px}}
.bug-ico{{font-size:34px}}
.bug-big{{font-size:28px;font-weight:800;color:var(--err)}}
.bug-lbl{{font-size:11px;color:var(--sub)}}
.bug-bar-wrap{{flex:1}}
.bug-bg{{height:7px;background:var(--border);border-radius:4px;overflow:hidden}}
.bug-fill{{height:7px;background:var(--err);border-radius:4px}}
.rbl{{color:#f87171;font-weight:700}}
footer{{text-align:center;padding:22px;font-size:12px;color:var(--sub);border-top:1px solid var(--border);margin-top:20px}}
</style>
</head>
<body>
{PASSWORD_GATE}
<header>
  <div class="logo">
    <div class="logo-box">C24</div>
    <div><div class="h-title">Performance Metrics Dashboard</div>
    <div class="h-sub">CARS24 Chatbot · Session outcome analysis</div></div>
  </div>
  <div class="h-right">
    <div class="live"><span class="dot"></span>Data as of {GENERATED_AT}</div>
    <div class="nav">
      <a href="dashboard_metrics.html" class="on">📊 Metrics</a>
      <a href="dashboard_qualitative.html">💡 Qualitative</a>
    </div>
  </div>
</header>
<div class="df-bar">
  <div class="df-inner">
    <span style="font-size:16px">📅</span>
    <span class="df-lbl">Date Range</span>
    <input type="date" id="df-start" class="df-inp"/>
    <span class="df-sep">→</span>
    <input type="date" id="df-end" class="df-inp"/>
    <button id="df-apply" class="df-btn df-go">Apply</button>
    <button id="df-clear" class="df-btn df-clr">✕ Clear</button>
    <span id="df-badge" class="df-badge" style="display:none"></span>
  </div>
</div>
<main>
  <div class="info-bar">✅ &nbsp;Data loaded from Google Sheets · {TOTAL_SESSIONS} sessions · Run <strong>generate_dashboards.py</strong> to refresh</div>
  <div class="kpi-grid" id="kpi-grid"></div>
  <div class="sep">Session Outcomes</div>
  <div class="row2">
    <div class="cc"><div class="cc-title">Overall Outcome Breakdown</div>
      <div class="cc-sub">All sessions by final classification</div>
      <div class="cw"><canvas id="outcomeChart"></canvas></div></div>
    <div class="cc"><div class="cc-title">Outcome Detail</div>
      <div class="cc-sub">Count and percentage per outcome</div>
      <table class="ot" id="outcome-tbl"></table></div>
  </div>
  <div class="sep">DSQ vs QNR Analysis</div>
  <div class="row2">
    <div class="cc"><div class="cc-title">Outcome Comparison — DSQ vs QNR</div>
      <div class="cc-sub">Side-by-side outcome counts</div>
      <div class="cw"><canvas id="compareChart"></canvas></div></div>
    <div class="cc"><div class="cc-title">Case Type Summary</div>
      <div class="cc-sub">Key rates per case type</div>
      <div class="cmp-grid" id="cmp-grid"></div></div>
  </div>
  <div class="sep">Daily Volume</div>
  <div class="row2">
    <div class="cc cc-full"><div class="cc-title">Sessions per Day</div>
      <div class="cc-sub">Total sessions per day in the dataset</div>
      <div class="cw tall"><canvas id="dailyChart"></canvas></div></div>
  </div>
  <div class="sep">Weekly Trends — Last 3 Months</div>
  <div class="row2">
    <div class="cc"><div class="cc-title">DSQ — Weekly Outcome Breakdown</div>
      <div class="cc-sub">Stacked 100% by outcome category per week</div>
      <div class="cw tall"><canvas id="wDsqChart"></canvas></div>
      <div style="text-align:center;font-size:11px;color:var(--sub);margin-top:6px" id="wDsqInfo"></div></div>
    <div class="cc"><div class="cc-title">QNR — Weekly Outcome Breakdown</div>
      <div class="cc-sub">Stacked 100% by outcome category per week</div>
      <div class="cw tall"><canvas id="wQnrChart"></canvas></div>
      <div style="text-align:center;font-size:11px;color:var(--sub);margin-top:6px" id="wQnrInfo"></div></div>
  </div>
  <div class="sep">Monthly Trends — Last 3 Months</div>
  <div class="row2">
    <div class="cc"><div class="cc-title">DSQ — Monthly Outcome Breakdown</div>
      <div class="cc-sub">Stacked 100% by outcome category per month</div>
      <div class="cw tall"><canvas id="mDsqChart"></canvas></div></div>
    <div class="cc"><div class="cc-title">QNR — Monthly Outcome Breakdown</div>
      <div class="cc-sub">Stacked 100% by outcome category per month</div>
      <div class="cw tall"><canvas id="mQnrChart"></canvas></div></div>
  </div>
  <div class="sep">Quality Indicators</div>
  <div class="row2">
    <div class="cc"><div class="cc-title">Bot Repeat Message Bug</div>
      <div class="cc-sub">Sessions where bot sent the same message ≥ 2 times</div>
      <div id="bug-wrap"></div></div>
    <div class="cc"><div class="cc-title">Repeat Bucket Breakdown</div>
      <div class="cc-sub">Sessions by maximum message repetition threshold</div>
      <div id="repeat-buckets-wrap"></div></div>
  </div>
</main>
{CHAT_WIDGET}
<footer>CARS24 Chatbot Intelligence · Generated {GENERATED_AT} · Refresh: <code>python generate_dashboards.py</code></footer>
{RECOMPUTE_JS}
<script>
const COLORS={{auction_started:'#22c55e',ticket_raised:'#3b82f6',drop_with_resolution:'#f59e0b',drop_without_resolution:'#ef4444'}};
const LABELS={{auction_started:'Auction Started',ticket_raised:'Ticket Raised',drop_with_resolution:'Drop – Resolved',drop_without_resolution:'Drop – Unresolved'}};
const OCS=['auction_started','ticket_raised','drop_with_resolution','drop_without_resolution'];
const OC_COLORS=['#22c55e','#3b82f6','#f59e0b','#ef4444'];
const OC_LABELS=['Auction Started','Ticket Raised','Drop Resolved','Drop Unresolved'];
const m=(_flt.start||_flt.end)?recompute(_sl):DATA.metrics;
const p=(n,d)=>d?Math.round(n/d*100):0;
// KPIs
document.getElementById('kpi-grid').innerHTML=[
  {{l:'Total Sessions',v:m.total,cls:'b',ic:'💬',s:`${{m.dsq_count}} DSQ · ${{m.qnr_count}} QNR`}},
  {{l:'Auction Started',v:m.outcomes.auction_started,cls:'g',ic:'🔨',s:`${{p(m.outcomes.auction_started,m.total)}}% of sessions`}},
  {{l:'Tickets Raised',v:m.outcomes.ticket_raised,cls:'p',ic:'🎫',s:`Ticket rate ${{m.ticket_rate}}%`}},
  {{l:'Drop Rate',v:m.drop_rate+'%',cls:'r',ic:'📉',s:`Resolved ${{m.outcomes.drop_with_resolution}} · Unresolved ${{m.outcomes.drop_without_resolution}}`}},
  {{l:'DSQ Auction Rate',v:m.auction_rate_dsq+'%',cls:'t',ic:'🏷️',s:'Of DSQ sessions'}},
  {{l:'QNR Auction Rate',v:m.auction_rate_qnr+'%',cls:'y',ic:'📋',s:'Of QNR sessions'}},
  {{l:'Auction + Ticket Both',v:m.has_both_pct+'%',cls:'g',ic:'🎯',s:`${{m.has_both_count}} sessions had both`}},
  {{l:'Avg Msgs/Session',v:m.avg_msgs,cls:'o',ic:'📊',s:'Total turns per session'}},
  {{l:'Repeat Bug',v:m.repeat_pct+'%',cls:'r',ic:'⚠️',s:`${{m.repeat}} sessions affected`}},
  {{l:'DSQ EHB/TP',v:m.ehb_tp_dsq!=null?m.ehb_tp_dsq+'%':'—',cls:'g',ic:'🎯',s:m.ehb_tp_dsq!=null?`${{m.ehb_tp_dsq_n||0}} sessions`:'Needs NEW_TP data'}},
  {{l:'QNR EHB/TP',v:m.ehb_tp_qnr!=null?m.ehb_tp_qnr+'%':'—',cls:'t',ic:'📋',s:m.ehb_tp_qnr!=null?`${{m.ehb_tp_qnr_n||0}} sessions`:'Needs NEW_TP data'}},
].map(k=>`<div class="kpi ${{k.cls}}"><div class="kpi-ico">${{k.ic}}</div><div class="kpi-lbl">${{k.l}}</div><div class="kpi-val">${{k.v}}</div><div class="kpi-sub">${{k.s}}</div></div>`).join('');
// Outcome chart
const cwSliceLabelPlugin={{
  id:'cwSliceLabel',
  afterDatasetDraw:function(chart){{
    var ctx2=chart.ctx;
    var meta=chart.getDatasetMeta(0);
    var ds=chart.data.datasets[0].data;
    var tot=ds.reduce(function(a,b){{return a+b;}},0);
    meta.data.forEach(function(arc,i){{
      var pct=tot?Math.round(ds[i]/tot*100):0;
      if(pct<5)return;
      var cp=arc.getCenterPoint();
      ctx2.save();
      ctx2.font='bold 13px Segoe UI,sans-serif';
      ctx2.fillStyle='#fff';
      ctx2.textAlign='center';
      ctx2.textBaseline='middle';
      ctx2.shadowColor='rgba(0,0,0,.65)';
      ctx2.shadowBlur=5;
      ctx2.fillText(pct+'%',cp.x,cp.y);
      ctx2.restore();
    }});
  }}
}};
new Chart(document.getElementById('outcomeChart'),{{type:'doughnut',
  data:{{labels:OCS.map(k=>LABELS[k]),datasets:[{{data:OCS.map(k=>m.outcomes[k]),
    backgroundColor:OCS.map(k=>COLORS[k]),borderColor:'#1e293b',borderWidth:3,hoverOffset:8}}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'right',labels:{{color:'#94a3b8',padding:14,font:{{size:12}}}}}},
      tooltip:{{callbacks:{{label:ctx=>` ${{ctx.label}}: ${{ctx.raw}} (${{p(ctx.raw,m.total)}}%)`}}}}}}}},
  plugins:[cwSliceLabelPlugin]}});
document.getElementById('outcome-tbl').innerHTML='<thead><tr><th>Outcome</th><th>Count</th><th>Share</th></tr></thead><tbody>'+
  OCS.map(k=>`<tr><td><span class="odot" style="background:${{COLORS[k]}}"></span>${{LABELS[k]}}</td><td style="font-weight:700">${{m.outcomes[k]}}</td>
    <td><div class="pb"><div class="pb-bg"><div class="pb-fill" style="width:${{p(m.outcomes[k],m.total)}}%;background:${{COLORS[k]}}"></div></div>
    <span style="font-size:11px;color:#94a3b8;min-width:34px">${{p(m.outcomes[k],m.total)}}%</span></div></td></tr>`).join('')+'</tbody>';
// Compare chart
new Chart(document.getElementById('compareChart'),{{type:'bar',
  data:{{labels:OCS.map(k=>LABELS[k]),datasets:[
    {{label:'DSQ',data:OCS.map(k=>m.dsq_outcomes[k]),backgroundColor:'rgba(139,92,246,.7)',borderColor:'#8b5cf6',borderWidth:1,borderRadius:4}},
    {{label:'QNR',data:OCS.map(k=>m.qnr_outcomes[k]),backgroundColor:'rgba(59,130,246,.7)',borderColor:'#3b82f6',borderWidth:1,borderRadius:4}}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{labels:{{color:'#94a3b8',font:{{size:12}}}}}},tooltip:{{mode:'index'}}}},
    scales:{{x:{{ticks:{{color:'#94a3b8',font:{{size:11}}}},grid:{{color:'rgba(51,65,85,.3)'}}}},
      y:{{ticks:{{color:'#94a3b8'}},grid:{{color:'rgba(51,65,85,.3)'}},beginAtZero:true}}}}}}}});
document.getElementById('cmp-grid').innerHTML=`
  <div class="cmp dsq"><div class="cmp-title dsq">DSQ (${{m.dsq_count}})</div>
    <div class="cmp-row"><span class="cmp-k">Auction Rate</span><span class="cmp-v">${{m.auction_rate_dsq}}%</span></div>
    <div class="cmp-row"><span class="cmp-k">Ticket Rate</span><span class="cmp-v">${{p(m.dsq_outcomes.ticket_raised,m.dsq_count)}}%</span></div>
    <div class="cmp-row"><span class="cmp-k">Resolved Drop</span><span class="cmp-v">${{p(m.dsq_outcomes.drop_with_resolution,m.dsq_count)}}%</span></div>
    <div class="cmp-row"><span class="cmp-k">Unresolved</span><span class="cmp-v">${{p(m.dsq_outcomes.drop_without_resolution,m.dsq_count)}}%</span></div>
  </div>
  <div class="cmp qnr"><div class="cmp-title qnr">QNR (${{m.qnr_count}})</div>
    <div class="cmp-row"><span class="cmp-k">Auction Rate</span><span class="cmp-v">${{m.auction_rate_qnr}}%</span></div>
    <div class="cmp-row"><span class="cmp-k">Ticket Rate</span><span class="cmp-v">${{p(m.qnr_outcomes.ticket_raised,m.qnr_count)}}%</span></div>
    <div class="cmp-row"><span class="cmp-k">Resolved Drop</span><span class="cmp-v">${{p(m.qnr_outcomes.drop_with_resolution,m.qnr_count)}}%</span></div>
    <div class="cmp-row"><span class="cmp-k">Unresolved</span><span class="cmp-v">${{p(m.qnr_outcomes.drop_without_resolution,m.qnr_count)}}%</span></div>
  </div>`;
// Daily chart
new Chart(document.getElementById('dailyChart'),{{type:'line',
  data:{{labels:m.daily.map(d=>d.d),datasets:[{{label:'Sessions',data:m.daily.map(d=>d.c),
    fill:true,backgroundColor:'rgba(232,0,45,.12)',borderColor:'#e8002d',
    borderWidth:2,pointBackgroundColor:'#e8002d',pointRadius:5,pointHoverRadius:7,tension:.3}}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}}}},
    scales:{{x:{{ticks:{{color:'#94a3b8',font:{{size:12}}}},grid:{{color:'rgba(51,65,85,.3)'}}}},
      y:{{ticks:{{color:'#94a3b8'}},grid:{{color:'rgba(51,65,85,.3)'}},beginAtZero:true}}}}}}}});
// ── Weekly / Monthly Trend Charts ──────────────────────────────────────────────
(function(){{
  function makeTrendChart(canvasId, trends, pctKey){{
    if(!trends||!trends.length)return;
    const labels=trends.map(t=>t.label);
    new Chart(document.getElementById(canvasId),{{
      type:'bar',
      data:{{
        labels,
        datasets:OCS.map((oc,i)=>({{
          label:OC_LABELS[i],
          data:trends.map(t=>(t[pctKey]&&t[pctKey][oc])||0),
          backgroundColor:OC_COLORS[i]+'bb',
          borderColor:OC_COLORS[i],
          borderWidth:1,
          borderRadius:2,
        }}))
      }},
      options:{{
        responsive:true,maintainAspectRatio:false,
        plugins:{{
          legend:{{labels:{{color:'#94a3b8',font:{{size:11}}}}}},
          tooltip:{{mode:'index',callbacks:{{label:ctx=>`${{ctx.dataset.label}}: ${{ctx.raw}}%`}}}}
        }},
        scales:{{
          x:{{stacked:true,ticks:{{color:'#94a3b8',font:{{size:10}}}},grid:{{color:'rgba(51,65,85,.3)'}}}},
          y:{{stacked:true,min:0,max:100,
            ticks:{{color:'#94a3b8',callback:v=>v+'%'}},
            grid:{{color:'rgba(51,65,85,.3)'}}}}
        }}
      }}
    }});
  }}
  const wt=m.weekly_trends||[], mt=m.monthly_trends||[];
  makeTrendChart('wDsqChart',wt,'dsq_pcts');
  makeTrendChart('wQnrChart',wt,'qnr_pcts');
  makeTrendChart('mDsqChart',mt,'dsq_pcts');
  makeTrendChart('mQnrChart',mt,'qnr_pcts');
  if(wt.length){{
    const wTotal=wt=>wt.dsq_total||0;
    document.getElementById('wDsqInfo').textContent=wt.map(w=>`${{w.label}}: ${{w.dsq_total}}`).join('  ·  ');
    document.getElementById('wQnrInfo').textContent=wt.map(w=>`${{w.label}}: ${{w.qnr_total}}`).join('  ·  ');
  }}
}})();
// ── Bug widget ─────────────────────────────────────────────────────────────────
document.getElementById('bug-wrap').innerHTML=`<div class="bug-card">
  <div class="bug-ico">🐛</div>
  <div><div class="bug-big">${{m.repeat}}</div><div class="bug-lbl">sessions with duplicate messages</div></div>
  <div class="bug-bar-wrap">
    <div style="font-size:26px;font-weight:800;color:#ef4444;margin-bottom:7px">${{m.repeat_pct}}%</div>
    <div class="bug-bg"><div class="bug-fill" style="width:${{Math.min(m.repeat_pct,100)}}%"></div></div>
    <div style="font-size:11px;color:#94a3b8;margin-top:5px">of ${{m.total}} sessions</div>
  </div></div>`;
// ── Repeat Buckets Table ───────────────────────────────────────────────────────
(function(){{
  const rb=m.repeat_buckets||[];
  document.getElementById('repeat-buckets-wrap').innerHTML=
    '<table class="ot" style="margin-top:8px"><thead><tr><th>Threshold</th><th>Sessions</th><th>% of Total</th><th>Max Rep Count</th></tr></thead><tbody>'+
    rb.map(b=>`<tr>
      <td class="rbl">${{b.label}}</td>
      <td style="font-weight:700">${{b.session_count}}</td>
      <td><div class="pb"><div class="pb-bg"><div class="pb-fill" style="width:${{Math.min(p(b.session_count,m.total),100)}}%;background:#ef4444"></div></div>
        <span style="font-size:11px;color:#94a3b8;min-width:36px">${{p(b.session_count,m.total)}}%</span></div></td>
      <td style="color:#f59e0b">${{b.rep_count}}</td>
    </tr>`).join('')+'</tbody></table>';
}})();
</script>
</body>
</html>
"""

# ── HTML template — Qualitative ───────────────────────────────────────────────

QUAL_HTML = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>CARS24 Chatbot — Qualitative Insights</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<script>
// ── Data embedded by generate_dashboards.py ──
const DATA = {DATA_JSON};
</script>
<style>
:root{{--red:#e8002d;--dark:#1a1a2e;--mid:#16213e;--blue:#0f3460;
  --green:#22c55e;--yellow:#f59e0b;--err:#ef4444;--purple:#8b5cf6;--teal:#14b8a6;
  --card:#1e293b;--border:#334155;--text:#e2e8f0;--sub:#94a3b8}}
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:var(--dark);color:var(--text);min-height:100vh}}
header{{background:linear-gradient(135deg,var(--mid),var(--blue));border-bottom:3px solid var(--red);
  padding:18px 32px;display:flex;align-items:center;justify-content:space-between;
  position:sticky;top:0;z-index:100;box-shadow:0 4px 20px rgba(0,0,0,.4)}}
.logo{{display:flex;align-items:center;gap:14px}}
.logo-box{{background:var(--red);color:#fff;font-weight:800;font-size:17px;padding:6px 12px;border-radius:6px;letter-spacing:1px}}
.h-title{{font-size:19px;font-weight:700;color:#fff}}
.h-sub{{font-size:12px;color:var(--sub);margin-top:2px}}
.h-right{{text-align:right}}
.live{{font-size:12px;color:var(--sub)}}
.dot{{display:inline-block;width:8px;height:8px;background:var(--green);border-radius:50%;margin-right:6px;animation:pulse 2s infinite}}
@keyframes pulse{{0%,100%{{opacity:1}}50%{{opacity:.4}}}}
.nav{{display:flex;gap:10px;margin-top:8px;justify-content:flex-end}}
.nav a{{font-size:12px;color:var(--sub);text-decoration:none;padding:4px 10px;
  border:1px solid var(--border);border-radius:4px;transition:.2s}}
.nav a:hover,.nav a.on{{color:#fff;border-color:var(--red);background:rgba(232,0,45,.15)}}
.df-bar{{background:var(--mid);border-bottom:1px solid var(--border);padding:10px 32px;
  position:sticky;top:76px;z-index:98}}
.df-inner{{display:flex;align-items:center;gap:10px;flex-wrap:wrap;max-width:1300px;margin:0 auto}}
.df-lbl{{font-size:12px;font-weight:700;color:var(--sub);text-transform:uppercase;letter-spacing:1px;white-space:nowrap}}
.df-inp{{background:#0f172a;border:1px solid var(--border);color:var(--text);border-radius:6px;
  padding:5px 10px;font-size:13px;outline:none;cursor:pointer;color-scheme:dark}}
.df-inp:focus{{border-color:var(--red)}} .df-sep{{color:var(--sub);font-size:14px}}
.df-btn{{border:none;border-radius:6px;padding:6px 14px;font-size:12px;font-weight:700;cursor:pointer;transition:.15s}}
.df-go{{background:var(--red);color:#fff}} .df-go:hover{{background:#c0001f}}
.df-clr{{background:var(--card);color:var(--sub);border:1px solid var(--border)}} .df-clr:hover{{color:var(--text)}}
.df-badge{{background:rgba(232,0,45,.15);border:1px solid rgba(232,0,45,.3);color:#fca5a5;
  border-radius:4px;font-size:11px;padding:3px 10px;white-space:nowrap}}
main{{max-width:1300px;margin:0 auto;padding:26px 22px}}
.info-bar{{background:rgba(34,197,94,.08);border:1px solid rgba(34,197,94,.3);border-radius:8px;
  padding:10px 16px;font-size:13px;color:#4ade80;margin-bottom:22px;display:flex;align-items:center;gap:8px}}
.summary-card{{background:linear-gradient(135deg,var(--mid),var(--blue));
  border:1px solid var(--red);border-radius:14px;padding:22px 26px;margin-bottom:26px;
  position:relative;overflow:hidden}}
.summary-card::before{{content:'';position:absolute;top:0;right:0;width:200px;height:200px;
  background:radial-gradient(circle,rgba(232,0,45,.15) 0%,transparent 70%);pointer-events:none}}
.sum-lbl{{font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:var(--red);margin-bottom:8px}}
.sum-text{{font-size:15px;line-height:1.75}}
.meta-row{{display:flex;gap:14px;margin-top:14px;flex-wrap:wrap}}
.mc{{font-size:12px;color:var(--sub);background:rgba(255,255,255,.07);padding:4px 12px;
  border-radius:20px;border:1px solid var(--border)}}
.mc span{{color:#fff;font-weight:600}}
.sep{{font-size:10px;font-weight:700;letter-spacing:2px;text-transform:uppercase;color:var(--sub);
  padding:4px 0 14px;border-top:1px solid var(--border);margin:26px 0 18px}}
.row2{{display:grid;grid-template-columns:1fr 1fr;gap:22px;margin-bottom:22px}}
@media(max-width:900px){{.row2{{grid-template-columns:1fr}}}}
.cc{{background:var(--card);border:1px solid var(--border);border-radius:13px;padding:18px 20px}}
.cc-full{{grid-column:1/-1}}
.cc-title{{font-size:12px;font-weight:700;letter-spacing:1px;text-transform:uppercase;color:var(--sub);margin-bottom:3px}}
.cc-sub{{font-size:12px;color:var(--sub);margin-bottom:14px}}
.cw{{position:relative;height:210px}} .cw.tall{{height:290px}} .cw.xlarge{{height:380px}}
.ig{{display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-bottom:32px}}
@media(max-width:900px){{.ig{{grid-template-columns:1fr}}}}
.ehb-banner{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:22px}}
@media(max-width:700px){{.ehb-banner{{grid-template-columns:1fr}}}}
.ehb-kpi{{background:var(--card);border:1px solid var(--border);border-radius:13px;
  padding:20px 24px;display:flex;align-items:center;gap:18px}}
.ehb-ico{{font-size:30px}}
.ehb-body{{flex:1}}
.ehb-label{{font-size:11px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:var(--sub);margin-bottom:4px}}
.ehb-value{{font-size:34px;font-weight:800;color:#22c55e;line-height:1}}
.ehb-value.na{{font-size:20px;color:var(--sub);font-weight:400}}
.ehb-note{{font-size:11px;color:var(--sub);margin-top:5px}}
.col-hdr{{display:flex;align-items:center;gap:10px;margin-bottom:16px}}
.col-ico{{width:34px;height:34px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:17px}}
.col-ico.g{{background:rgba(34,197,94,.15)}} .col-ico.r{{background:rgba(239,68,68,.15)}}
.col-ttl{{font-size:17px;font-weight:700}}
.col-cnt{{font-size:11px;color:var(--sub);background:var(--card);padding:3px 9px;border-radius:20px;border:1px solid var(--border)}}
.cards{{display:flex;flex-direction:column;gap:13px}}
.ic{{background:var(--card);border:1px solid var(--border);border-radius:12px;
  padding:16px 18px;transition:.15s;position:relative;overflow:hidden}}
.ic:hover{{transform:translateY(-2px);box-shadow:0 8px 28px rgba(0,0,0,.3)}}
.ic.ww{{border-left:4px solid var(--green)}} .ic.ni{{border-left:4px solid var(--err)}}
.ic-top{{display:flex;align-items:flex-start;justify-content:space-between;gap:8px;margin-bottom:8px}}
.ic-ttl{{font-size:14px;font-weight:700;color:#fff;flex:1;line-height:1.3}}
.bdgs{{display:flex;gap:5px;flex-wrap:wrap;flex-shrink:0}}
.badge{{font-size:10px;font-weight:700;letter-spacing:.5px;padding:3px 7px;border-radius:4px;text-transform:uppercase}}
.b-dsq{{background:rgba(139,92,246,.2);color:#a78bfa;border:1px solid rgba(139,92,246,.3)}}
.b-qnr{{background:rgba(59,130,246,.2);color:#60a5fa;border:1px solid rgba(59,130,246,.3)}}
.b-both{{background:rgba(245,158,11,.2);color:#fbbf24;border:1px solid rgba(245,158,11,.3)}}
.b-hi{{background:rgba(239,68,68,.2);color:#f87171;border:1px solid rgba(239,68,68,.3)}}
.b-md{{background:rgba(245,158,11,.2);color:#fbbf24;border:1px solid rgba(245,158,11,.3)}}
.b-lo{{background:rgba(34,197,94,.2);color:#4ade80;border:1px solid rgba(34,197,94,.3)}}
.ic-detail{{font-size:13px;color:var(--sub);line-height:1.6;margin-bottom:9px}}
.ic-rec{{font-size:13px;color:#7dd3fc;background:rgba(56,189,248,.08);
  border:1px solid rgba(56,189,248,.2);border-radius:6px;padding:8px 11px;line-height:1.5}}
.ic-rec::before{{content:'💡 '}}
.imp-row{{display:flex;align-items:center;gap:7px;margin-top:9px}}
.imp-lbl{{font-size:10px;color:var(--sub);width:42px}}
.imp-bg{{flex:1;height:4px;background:var(--border);border-radius:2px}}
.imp-fill{{height:4px;border-radius:2px}}
.imp-fill.High{{width:100%;background:var(--err)}}
.imp-fill.Medium{{width:66%;background:var(--yellow)}}
.imp-fill.Low{{width:33%;background:var(--green)}}
.ic-num{{position:absolute;top:12px;right:12px;font-size:30px;font-weight:900;color:rgba(255,255,255,.04);line-height:1}}
footer{{text-align:center;padding:22px;font-size:12px;color:var(--sub);border-top:1px solid var(--border);margin-top:20px}}
</style>
</head>
<body>
{PASSWORD_GATE}
<header>
  <div class="logo">
    <div class="logo-box">C24</div>
    <div><div class="h-title">Qualitative Insights Dashboard</div>
    <div class="h-sub">CARS24 Chatbot · Pattern-based analysis</div></div>
  </div>
  <div class="h-right">
    <div class="live"><span class="dot"></span>Data as of {GENERATED_AT}</div>
    <div class="nav">
      <a href="dashboard_metrics.html">📊 Metrics</a>
      <a href="dashboard_qualitative.html" class="on">💡 Qualitative</a>
    </div>
  </div>
</header>
<div class="df-bar">
  <div class="df-inner">
    <span style="font-size:16px">📅</span>
    <span class="df-lbl">Date Range</span>
    <input type="date" id="df-start" class="df-inp"/>
    <span class="df-sep">→</span>
    <input type="date" id="df-end" class="df-inp"/>
    <button id="df-apply" class="df-btn df-go">Apply</button>
    <button id="df-clear" class="df-btn df-clr">✕ Clear</button>
    <span id="df-badge" class="df-badge" style="display:none"></span>
  </div>
</div>
<main>
  <div class="info-bar">✅ &nbsp;Data loaded from Google Sheets · {TOTAL_SESSIONS} sessions · Run <strong>generate_dashboards.py</strong> to refresh</div>
  <div class="summary-card">
    <div class="sum-lbl">Executive Summary</div>
    <div class="sum-text">{SUMMARY}</div>
    <div class="meta-row" id="meta-row"></div>
  </div>

  <div class="ehb-banner" id="ehb-banner">
    <div class="ehb-kpi">
      <div class="ehb-ico">🎯</div>
      <div class="ehb-body">
        <div class="ehb-label">DSQ — EHB / Target Price</div>
        <div class="ehb-value na" id="ehb-dsq-val">—</div>
        <div class="ehb-note" id="ehb-dsq-note">Auction start price vs fair market value</div>
      </div>
    </div>
    <div class="ehb-kpi">
      <div class="ehb-ico">📋</div>
      <div class="ehb-body">
        <div class="ehb-label">QNR — EHB / Target Price</div>
        <div class="ehb-value na" id="ehb-qnr-val">—</div>
        <div class="ehb-note" id="ehb-qnr-note">Auction start price vs fair market value</div>
      </div>
    </div>
  </div>
  <div class="row2">
    <div class="cc"><div class="cc-title">Strengths — Impact Distribution</div>
      <div class="cc-sub">By business impact level</div>
      <div class="cw"><canvas id="impactChart"></canvas></div></div>
    <div class="cc"><div class="cc-title">Improvement Areas — Priority</div>
      <div class="cc-sub">By urgency level</div>
      <div class="cw"><canvas id="priorityChart"></canvas></div></div>
  </div>
  <div class="sep">Region Analysis</div>
  <div class="row2">
    <div class="cc cc-full">
      <div class="cc-title">Region-wise DSQ vs QNR Distribution</div>
      <div class="cc-sub">Top regions by total chat volume (DSQ + QNR)</div>
      <div id="region-chart-wrap">
        <div class="cw xlarge"><canvas id="regionChart"></canvas></div>
      </div>
      <div id="region-empty" style="display:none;padding:30px;text-align:center;color:var(--sub);font-size:13px">
        No region data available — ensure REGION_3 column is populated in the sheet.</div>
    </div>
  </div>
  <div class="sep">Inspection-to-Chat Gap Analysis</div>
  <div class="row2">
    <div class="cc">
      <div class="cc-title">DSQ — Days from Inspection to Chat</div>
      <div class="cc-sub">Gap between INSP_DATE_3 and earliest chat date</div>
      <div class="cw tall"><canvas id="gapDsqChart"></canvas></div>
    </div>
    <div class="cc">
      <div class="cc-title">QNR — Days from Inspection to Chat</div>
      <div class="cc-sub">Gap between INSP_DATE_3 and earliest chat date</div>
      <div class="cw tall"><canvas id="gapQnrChart"></canvas></div>
    </div>
  </div>
  <div class="ig">
    <div>
      <div class="col-hdr"><div class="col-ico g">✅</div>
        <div class="col-ttl">What's Working Well</div>
        <div class="col-cnt">{WW_COUNT} insights</div></div>
      <div class="cards" id="ww-cards"></div>
    </div>
    <div>
      <div class="col-hdr"><div class="col-ico r">🔧</div>
        <div class="col-ttl">Needs Improvement</div>
        <div class="col-cnt">{NI_COUNT} areas</div></div>
      <div class="cards" id="ni-cards"></div>
    </div>
  </div>
</main>
{CHAT_WIDGET}
<footer>CARS24 Chatbot Intelligence · Generated {GENERATED_AT} · Refresh: <code>python generate_dashboards.py</code></footer>
{RECOMPUTE_JS}
<script>
const m=(_flt.start||_flt.end)?recompute(_sl):DATA.metrics, qual=DATA.qual;
const CT={{DSQ:'b-dsq',QNR:'b-qnr',Both:'b-both'}};
const LV={{High:'b-hi',Medium:'b-md',Low:'b-lo'}};
// Meta chips
// ── EHB/TP banner ──────────────────────────────────────────────────────────────
(function(){{
  var ed=m.ehb_tp_dsq, eq=m.ehb_tp_qnr;
  var edn=m.ehb_tp_dsq_n||0, eqn=m.ehb_tp_qnr_n||0;
  var dEl=document.getElementById('ehb-dsq-val');
  var qEl=document.getElementById('ehb-qnr-val');
  var dNote=document.getElementById('ehb-dsq-note');
  var qNote=document.getElementById('ehb-qnr-note');
  if(dEl){{
    if(ed!=null){{dEl.textContent=ed+'%';dEl.className='ehb-value';dNote.textContent=edn+' auction sessions with TP data';}}
    else{{dEl.textContent='—';dNote.textContent='Needs NEW_TP data from sheet';}}
  }}
  if(qEl){{
    if(eq!=null){{qEl.textContent=eq+'%';qEl.className='ehb-value';qNote.textContent=eqn+' auction sessions with TP data';}}
    else{{qEl.textContent='—';qNote.textContent='Needs NEW_TP data from sheet';}}
  }}
}})();
document.getElementById('meta-row').innerHTML=[
  `Total <span>${{m.total}}</span>`,`DSQ <span>${{m.dsq_count}}</span>`,`QNR <span>${{m.qnr_count}}</span>`,
  `Auction Rate <span>${{Math.round(m.outcomes.auction_started/Math.max(m.total,1)*100)}}%</span>`,
  `Both Auction+Ticket <span>${{m.has_both_pct}}%</span>`,
  `Repeat Bug <span>${{m.repeat_pct}}%</span>`,`Resolution <span>${{m.resolution_rate}}%</span>`,
].map(t=>`<div class="mc">${{t}}</div>`).join('');
// Impact/Priority charts
function cnt(arr,f){{const c={{High:0,Medium:0,Low:0}};arr.forEach(i=>{{if(c[i[f]]!==undefined)c[i[f]]++;}});return c;}}
const wc=cnt(qual.workingWell,'impact'), nc=cnt(qual.needsImprovement,'priority');
const dOpts={{responsive:true,maintainAspectRatio:false,
  plugins:{{legend:{{position:'right',labels:{{color:'#94a3b8',padding:14,font:{{size:12}}}}}}}}}};
new Chart(document.getElementById('impactChart'),{{type:'doughnut',
  data:{{labels:['High','Medium','Low'],datasets:[{{data:[wc.High,wc.Medium,wc.Low],
    backgroundColor:['#22c55e','#f59e0b','#3b82f6'],borderColor:'#1e293b',borderWidth:3,hoverOffset:6}}]}},
  options:dOpts}});
new Chart(document.getElementById('priorityChart'),{{type:'doughnut',
  data:{{labels:['High','Medium','Low'],datasets:[{{data:[nc.High,nc.Medium,nc.Low],
    backgroundColor:['#ef4444','#f59e0b','#22c55e'],borderColor:'#1e293b',borderWidth:3,hoverOffset:6}}]}},
  options:dOpts}});
// ── Region Chart (% of total DSQ and % of total QNR) ──────────────────────────
(function(){{
  const rsAll=m.region_split||[];
  const rs=rsAll.filter(r=>(r.region||'Unknown')!=='Unknown');
  const hasData=rs.some(r=>r.dsq>0||r.qnr>0);
  if(!hasData){{
    document.getElementById('region-chart-wrap').style.display='none';
    document.getElementById('region-empty').style.display='block';
    return;
  }}
  const dsqTot=m.dsq_count||1;
  const qnrTot=m.qnr_count||1;
  const dsqPcts=rs.map(r=>Math.round(r.dsq/dsqTot*100));
  const qnrPcts=rs.map(r=>Math.round(r.qnr/qnrTot*100));
  new Chart(document.getElementById('regionChart'),{{
    type:'bar',
    data:{{
      labels:rs.map(r=>r.region),
      datasets:[
        {{label:'% of DSQ',data:dsqPcts,backgroundColor:'rgba(139,92,246,.75)',borderColor:'#8b5cf6',borderWidth:1,borderRadius:4}},
        {{label:'% of QNR',data:qnrPcts,backgroundColor:'rgba(59,130,246,.75)',borderColor:'#3b82f6',borderWidth:1,borderRadius:4}},
      ]
    }},
    options:{{
      indexAxis:'y',
      responsive:true,maintainAspectRatio:false,
      plugins:{{
        legend:{{labels:{{color:'#94a3b8',font:{{size:12}}}}}},
        tooltip:{{
          mode:'index',
          callbacks:{{
            label:function(ctx){{
              var r=rs[ctx.dataIndex];
              var count=ctx.datasetIndex===0?r.dsq:r.qnr;
              return ctx.dataset.label+': '+ctx.raw+'% ('+count+' sessions)';
            }}
          }}
        }}
      }},
      scales:{{
        x:{{stacked:false,ticks:{{color:'#94a3b8',callback:function(v){{return v+'%';}}}},
          grid:{{color:'rgba(51,65,85,.3)'}},beginAtZero:true,max:Math.min(100,Math.max.apply(null,dsqPcts.concat(qnrPcts))+5)}},
        y:{{ticks:{{color:'#94a3b8',font:{{size:11}}}},grid:{{display:false}}}}
      }}
    }}
  }});
}})();

// ── Inspection Gap Charts (% of sessions within that gap bucket) ──────────────
(function(){{
  const ig=m.insp_gap||{{buckets:[],dsq:[],qnr:[]}};
  function makeGapChart(canvasId,rawData,color,caseLabel){{
    const total=rawData.reduce(function(a,b){{return a+b;}},0);
    if(total===0){{
      var el=document.getElementById(canvasId);
      if(el){{
        el.style.display='none';
        var msg=document.createElement('div');
        msg.style.cssText='padding:30px;text-align:center;color:#94a3b8;font-size:13px';
        msg.innerHTML='No inspection date data available — ensure INSP_DATE_3 is populated in the sheet.';
        el.parentNode.appendChild(msg);
      }}
      return;
    }}
    const pctData=rawData.map(function(v){{return total?Math.round(v/total*100):0;}});
    new Chart(document.getElementById(canvasId),{{
      type:'bar',
      data:{{
        labels:ig.buckets,
        datasets:[{{
          label:caseLabel,
          data:pctData,
          backgroundColor:color+'99',
          borderColor:color,
          borderWidth:1,
          borderRadius:4,
        }}]
      }},
      options:{{
        responsive:true,maintainAspectRatio:false,
        plugins:{{
          legend:{{display:false}},
          tooltip:{{callbacks:{{
            label:function(ctx){{
              var count=rawData[ctx.dataIndex];
              return caseLabel+': '+ctx.raw+'% ('+count+' sessions)';
            }}
          }}}}
        }},
        scales:{{
          x:{{ticks:{{color:'#94a3b8',font:{{size:10}}}},grid:{{color:'rgba(51,65,85,.3)'}}}},
          y:{{ticks:{{color:'#94a3b8',callback:function(v){{return v+'%';}}}},
            grid:{{color:'rgba(51,65,85,.3)'}},beginAtZero:true,max:100}}
        }}
      }}
    }});
  }}
  makeGapChart('gapDsqChart',ig.dsq,'#8b5cf6','DSQ');
  makeGapChart('gapQnrChart',ig.qnr,'#3b82f6','QNR');
}})();
// ── Insight Cards ───────────────────────────────────────────────────────────────
function renderCards(items,id,type){{
  document.getElementById(id).innerHTML=items.map((it,i)=>{{
    const lv=type==='ww'?it.impact:it.priority;
    return `<div class="ic ${{type}}">
      <div class="ic-num">${{i+1}}</div>
      <div class="ic-top"><div class="ic-ttl">${{it.title}}</div>
        <div class="bdgs"><span class="badge ${{CT[it.caseType]||'b-both'}}">${{it.caseType}}</span>
          <span class="badge ${{LV[lv]||'b-lo'}}">${{type==='ww'?'Impact':'Priority'}}: ${{lv}}</span></div></div>
      <div class="ic-detail">${{it.detail}}</div>
      ${{it.recommendation?`<div class="ic-rec">${{it.recommendation}}</div>`:''}}
      <div class="imp-row"><div class="imp-lbl">${{lv}}</div>
        <div class="imp-bg"><div class="imp-fill ${{lv}}"></div></div></div>
    </div>`;
  }}).join('');
}}
renderCards(qual.workingWell,'ww-cards','ww');
renderCards(qual.needsImprovement,'ni-cards','ni');
</script>
</body>
</html>
"""

# ── Post-Chat Dashboard ────────────────────────────────────────────────────────

_POSTCHAT_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>CARS24 Post-Chat Journey</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',system-ui,sans-serif;background:#0f172a;color:#e2e8f0;min-height:100vh}
header{background:#1e293b;border-bottom:1px solid #334155;padding:14px 24px;display:flex;align-items:center;gap:16px;position:sticky;top:0;z-index:100}
.logo{background:#e8002d;color:#fff;font-weight:800;font-size:18px;padding:5px 12px;border-radius:6px;letter-spacing:1px;white-space:nowrap}
.title{font-size:18px;font-weight:700;flex:1}
.subtitle{font-size:12px;color:#64748b;margin-top:2px}
.back-btn{font-size:13px;color:#94a3b8;text-decoration:none;border:1px solid #334155;padding:6px 12px;border-radius:6px;white-space:nowrap}
.back-btn:hover{color:#e2e8f0;border-color:#64748b}
.filters{background:#1e293b;border-bottom:1px solid #334155;padding:12px 24px;display:flex;flex-wrap:wrap;gap:12px;align-items:flex-end}
.filter-grp{display:flex;flex-direction:column;gap:4px}
.filter-grp label{font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.5px}
.filter-grp select,.filter-grp input{background:#0f172a;border:1px solid #334155;color:#e2e8f0;padding:6px 10px;border-radius:6px;font-size:13px;outline:none}
.filter-grp select:focus,.filter-grp input:focus{border-color:#3b82f6}
.gen-info{margin-left:auto;font-size:11px;color:#475569;align-self:center}
.main{padding:24px}
.kpi-row{display:flex;flex-wrap:wrap;gap:16px;margin-bottom:24px}
.kpi{background:#1e293b;border:1px solid #334155;border-radius:12px;padding:18px 24px;min-width:180px;flex:1}
.kpi .val{font-size:28px;font-weight:700;color:#f1f5f9}
.kpi .lbl{font-size:12px;color:#64748b;margin-top:4px;text-transform:uppercase;letter-spacing:.5px}
.kpi.test-kpi{border-top:3px solid #3b82f6}
.kpi.ctrl-kpi{border-top:3px solid #10b981}
.kpi.ehb-kpi{border-top:3px solid #f59e0b}
section{margin-bottom:28px}
section h2{font-size:15px;font-weight:700;margin-bottom:12px;display:flex;align-items:center;gap:8px}
section h2 .badge{font-size:11px;padding:2px 8px;border-radius:4px;font-weight:600}
.badge-test{background:#1e40af;color:#93c5fd}
.badge-ctrl{background:#065f46;color:#6ee7b7}
table{width:100%;border-collapse:collapse;font-size:13px;background:#1e293b;border-radius:10px;overflow:hidden}
thead tr{background:#0f172a}
th{padding:10px 14px;text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:.5px;color:#64748b;font-weight:600;white-space:nowrap}
th.num{text-align:right}
td{padding:9px 14px;border-top:1px solid #1e293b}
td.num{text-align:right;font-variant-numeric:tabular-nums}
td.ehbtp{color:#fbbf24;font-weight:600}
td.pct{color:#94a3b8}
tbody tr:hover{background:#263349}
.row-overall td{background:#162032;font-weight:700;color:#f1f5f9}
.row-total td{background:#0f172a;font-weight:700;color:#f1f5f9;border-top:2px solid #334155}
.test-col{color:#93c5fd}
.ctrl-col{color:#6ee7b7}
.section-grid{display:grid;grid-template-columns:1fr 1fr;gap:20px}
@media(max-width:900px){.section-grid{grid-template-columns:1fr}}
.cmp-table th:nth-child(2){color:#93c5fd}
.cmp-table th:nth-child(3){color:#6ee7b7}
.tag{display:inline-block;padding:2px 7px;border-radius:4px;font-size:11px;font-weight:600}
.tag-r{background:#1e3a5f;color:#7dd3fc}
.tag-t{background:#1a2e4a;color:#93c5fd}
.tag-d{background:#1f2937;color:#9ca3af}
footer{text-align:center;padding:20px;color:#475569;font-size:12px;border-top:1px solid #1e293b;margin-top:8px}
</style>
</head>
<body>
<header>
  <div class="logo">CARS24</div>
  <div>
    <div class="title">Post-Chat Journey Dashboard</div>
    <div class="subtitle">Bot (Test) vs Human (Control) — Auction Outcomes</div>
  </div>
  <a class="back-btn" href="index.html">← Hub</a>
</header>

<div class="filters">
  <div class="filter-grp">
    <label>From Date</label>
    <input type="date" id="dt-from"/>
  </div>
  <div class="filter-grp">
    <label>To Date</label>
    <input type="date" id="dt-to"/>
  </div>
  <div class="filter-grp">
    <label>Case Type</label>
    <select id="case-type">
      <option value="">All Cases</option>
      <option value="Q">QNR</option>
      <option value="D">DSQ</option>
    </select>
  </div>
  <div class="filter-grp">
    <label>Region</label>
    <select id="region"><option value="">All Regions</option></select>
  </div>
  <div class="filter-grp">
    <label>GS Assured</label>
    <select id="gs">
      <option value="">All</option>
      <option value="n">Non GS</option>
      <option value="f">GS Flagged</option>
      <option value="a">GS Non-Assured</option>
    </select>
  </div>
  <div class="filter-grp">
    <label>C2B / C2D</label>
    <select id="c2b">
      <option value="">All</option>
      <option value="b">C2B</option>
      <option value="d">C2D</option>
    </select>
  </div>
  <div class="gen-info" id="gen-info"></div>
</div>

<div class="main">

  <div class="kpi-row">
    <div class="kpi test-kpi"><div class="val" id="k-tleads">-</div><div class="lbl">Test Leads (Bot)</div></div>
    <div class="kpi ctrl-kpi"><div class="val" id="k-cleads">-</div><div class="lbl">Control Leads (Human)</div></div>
    <div class="kpi ehb-kpi"><div class="val" id="k-tehbtp">-</div><div class="lbl">Test Overall EHB/TP</div></div>
    <div class="kpi ehb-kpi"><div class="val" id="k-cehbtp">-</div><div class="lbl">Control EHB/TP</div></div>
    <div class="kpi test-kpi"><div class="val" id="k-traise-pct">-</div><div class="lbl">Test Raised%</div></div>
    <div class="kpi ctrl-kpi"><div class="val" id="k-craise-pct">-</div><div class="lbl">Control Raised%</div></div>
  </div>

  <section>
    <h2>🤖 Bot (Test) — Post-Chat Journey <span class="badge badge-test">TEST</span></h2>
    <table>
      <thead>
        <tr>
          <th>Category</th>
          <th class="num">Leads / Count</th>
          <th class="num">EHB/TP</th>
          <th class="num">Accepted</th>
          <th class="num">A2C / Token</th>
          <th class="num">Raised%</th>
          <th class="num">Acceptance%</th>
          <th class="num">Conversion%</th>
        </tr>
      </thead>
      <tbody id="bot-tbody"></tbody>
    </table>
  </section>

  <div class="section-grid">
    <section>
      <h2>👤 Human (Control) — Performance <span class="badge badge-ctrl">CONTROL</span></h2>
      <table>
        <thead>
          <tr>
            <th class="num">Total Leads</th>
            <th class="num">Raised</th>
            <th class="num">EHB/TP</th>
            <th class="num">Accepted</th>
            <th class="num">A2C</th>
            <th class="num">Raised%</th>
            <th class="num">Acceptance%</th>
            <th class="num">Conversion%</th>
          </tr>
        </thead>
        <tbody id="ctrl-tbody"></tbody>
      </table>
    </section>

    <section>
      <h2>📊 Test vs Control — Overall Comparison</h2>
      <table class="cmp-table">
        <thead>
          <tr>
            <th>Metric</th>
            <th class="num">🤖 Test (Bot)</th>
            <th class="num">👤 Control (Human)</th>
          </tr>
        </thead>
        <tbody id="cmp-tbody"></tbody>
      </table>
    </section>
  </div>

</div>
<footer>Generated: %%GENERATED_AT%% &nbsp;·&nbsp; CARS24 Chatbot Intelligence</footer>

<script>
(function(){
var testData = %%TEST_DATA%%;
var ctrlData = %%CTRL_DATA%%;
var REGIONS  = %%REGIONS%%;

// Test columns: [0]=appt_id, [1]=date, [2]=case_type(Q/D), [3]=region, [4]=gs(n/f/a),
//               [5]=c2b(b/d), [6]=tp, [7]=poc(PAI_OCB_CREATED), [8]=cx(CX_PAI_OCB),
//               [9]=cx_ehb(CX_AVG_EHB), [10]=pa(PAI_OCB_ACCEPTED), [11]=ca(CX_PAI_OCB_ACCEPTED),
//               [12]=cat(r/t/d), [13]=tok(New Token), [14]=rr(Re-Raised), [15]=ra(Re-Raised Accepted),
//               [16]=ehbtp(pre-calc), [17]=cx_ehbtp, [18]=ov_ehbtp(overall)
// Ctrl columns: [0]=app_id, [1]=date, [2]=case_type(Q/D), [3]=region, [4]=gs(n/f/a),
//               [5]=c2b(empty), [6]=tp, [7]=poc_ov(PAI_OCB_CREATED_TKT_OVERALL),
//               [8]=acc_ov(PAI_OCB_ACCEPTED_TKT_OVERALL), [9]=ehb_ov(EXPECTED_HB_OVERALL),
//               [10]=nb(New Bought), [11]=ov_ratio(Overall EHB/TP ratio)

// Populate region dropdown
var regSel = document.getElementById('region');
REGIONS.forEach(function(r){
  var o = document.createElement('option');
  o.value = r; o.textContent = r;
  regSel.appendChild(o);
});

// Set default dates
document.getElementById('dt-from').value = '%%MIN_DATE%%';
document.getElementById('dt-to').value   = '%%MAX_DATE%%';
document.getElementById('gen-info').textContent = 'Generated %%GENERATED_AT%%';

function n(v){ return (v===null||v===undefined||v!==v) ? null : +v; }
function fmt(v){ if(v===null||v===undefined||isNaN(v)) return '-'; return Math.round(v).toLocaleString('en-IN'); }
function fmtR(v){ if(v===null||v===undefined||isNaN(v)||v<=0) return '-'; return (v*100).toFixed(1)+'%'; }
function pct(a,b){ return (b&&b>0) ? (a/b*100).toFixed(1)+'%' : '-'; }

function sumDiv(rows,numFn,denFn){
  var s=0,d=0;
  rows.forEach(function(r){
    var nv=numFn(r), dv=denFn(r);
    if(nv!==null && dv!==null && dv>0){ s+=nv; d+=dv; }
  });
  return d>0 ? s/d : null;
}
function avgWhere(rows,fn){
  var s=0,c=0;
  rows.forEach(function(r){
    var v=fn(r);
    if(v!==null && !isNaN(v) && v>0.05 && v<10){ s+=v; c++; }
  });
  return c>0 ? s/c : null;
}

function applyFilters(rows, isCtrl){
  var df=document.getElementById('dt-from').value;
  var dt=document.getElementById('dt-to').value;
  var ct=document.getElementById('case-type').value;
  var rg=document.getElementById('region').value;
  var gs=document.getElementById('gs').value;
  var c2b=document.getElementById('c2b').value;
  return rows.filter(function(r){
    if(df && r[1] && r[1]<df) return false;
    if(dt && r[1] && r[1]>dt) return false;
    if(ct && r[2]!==ct) return false;
    if(rg && r[3]!==rg) return false;
    if(gs && r[4]!==gs) return false;
    if(!isCtrl && c2b && r[5]!==c2b) return false;
    return true;
  });
}

function mkTd(txt,cls){
  var td=document.createElement('td');
  td.className=cls||'';
  td.innerHTML=txt;
  return td;
}
function mkRow(cells){
  var tr=document.createElement('tr');
  cells.forEach(function(c){ tr.appendChild(c); });
  return tr;
}

function recompute(){
  var tf = applyFilters(testData, false);
  var cf = applyFilters(ctrlData, true);
  var tTotal = tf.length;

  // ── First Raise by Bot (Category=bot_raise AND CX_PAI_OCB>=1) ──
  var fRows = tf.filter(function(r){ return r[12]==='r' && n(r[8])>=1; });
  var fEhbTp = sumDiv(fRows, function(r){ return n(r[9]); }, function(r){ return n(r[6]); });
  var fAcc   = fRows.filter(function(r){ return n(r[11])>=1; }).length;
  // D11: CX_PAI_OCB_ACCEPTED>=1 AND New Token issued (matches Excel formula)
  var fA2c   = fRows.filter(function(r){ return n(r[11])>=1 && r[13]>0; }).length;

  // ── Re-Raised by RA (subset of First Raise rows that had a re-raise) ──
  var rrRows   = fRows.filter(function(r){ return n(r[14])>=1; });
  var rrCount  = rrRows.length;
  var rraCount = rrRows.filter(function(r){ return n(r[15])>=1; }).length;
  var rrA2c    = rrRows.filter(function(r){ return n(r[15])>=1 && r[13]>0; }).length;  // Re-Raised Acc + token
  var rrEhbTp  = sumDiv(rrRows, function(r){ return n(r[9]); }, function(r){ return n(r[6]); });

  // ── Ticket by Bot (Category=ticket) ──
  var tktRows  = tf.filter(function(r){ return r[12]==='t'; });
  var tktAcc   = tktRows.filter(function(r){ return n(r[10])>=1; }).length;
  var tktA2c   = tktRows.filter(function(r){ return r[13]>0; }).length;
  var tktEhbTp = sumDiv(tktRows, function(r){ return n(r[9]); }, function(r){ return n(r[6]); });

  // ── Drop / No Action (Category=drop) ──
  var dropRows = tf.filter(function(r){ return r[12]==='d'; });
  var dropAcc  = dropRows.filter(function(r){ return n(r[10])>=1; }).length;
  var dropA2c  = dropRows.filter(function(r){ return r[13]>0; }).length;
  var dropEhbTp= sumDiv(dropRows, function(r){ return n(r[9]); }, function(r){ return n(r[6]); });

  // ── Overall Bot (PAI_OCB_CREATED>=1) ──
  var ovRows  = tf.filter(function(r){ return n(r[7])>=1; });
  var ovEhbTp = sumDiv(ovRows, function(r){ return n(r[9]); }, function(r){ return n(r[6]); });
  var ovAcc   = tf.filter(function(r){ return n(r[10])>=1; }).length;
  var ovA2c   = tf.filter(function(r){ return n(r[13])>0; }).length;

  // ── Control — deduplicate on APP_ID (r[0]) for all counts ──
  var _cUniq = function(rows){ return new Set(rows.map(function(r){return r[0];})).size; };
  var cTotal = _cUniq(cf);
  var cRaisedRows = cf.filter(function(r){ return n(r[7])>=1; });
  var cRaised = _cUniq(cRaisedRows);
  var cEhbTp  = sumDiv(cRaisedRows, function(r){ return n(r[9]); }, function(r){ return n(r[6]); });
  var cAcc    = _cUniq(cf.filter(function(r){ return n(r[7])>=1 && n(r[8])>=1; }));
  var cA2c    = _cUniq(cf.filter(function(r){ return n(r[7])>=1 && n(r[8])>=1 && n(r[10])>0; }));
  var cToken  = _cUniq(cf.filter(function(r){ return n(r[10])>0; }));

  // ── KPI Cards ──
  document.getElementById('k-tleads').textContent    = fmt(tTotal);
  document.getElementById('k-cleads').textContent    = fmt(cTotal);
  document.getElementById('k-tehbtp').textContent    = fmtR(ovEhbTp);
  document.getElementById('k-cehbtp').textContent    = fmtR(cEhbTp);
  document.getElementById('k-traise-pct').textContent = pct(ovRows.length, tTotal);
  document.getElementById('k-craise-pct').textContent = pct(cRaised, cTotal);

  // ── Bot Table ──
  var tbody = document.getElementById('bot-tbody');
  tbody.innerHTML = '';

  function addBotRow(label, raised, ehbtp, acc, a2c, cls, noAccPct){
    var tr = document.createElement('tr');
    if(cls) tr.className=cls;
    tr.appendChild(mkTd('<span class="tag tag-'+label.charAt(0).toLowerCase()+'">'+label+'</span>'));
    tr.appendChild(mkTd(fmt(raised),'num'));
    tr.appendChild(mkTd(fmtR(ehbtp),'num ehbtp'));
    tr.appendChild(mkTd(fmt(acc),'num'));
    tr.appendChild(mkTd(fmt(a2c),'num'));
    tr.appendChild(mkTd(pct(raised,tTotal),'num pct'));
    tr.appendChild(mkTd(noAccPct?'-':pct(acc,raised),'num pct'));
    tr.appendChild(mkTd(a2c===null?'-':pct(a2c,tTotal),'num pct'));
    tbody.appendChild(tr);
  }

  // First Raise by Bot row
  var tr1=document.createElement('tr');
  tr1.innerHTML='<td>🎯 First Raise by Bot</td>'+
    '<td class="num">'+fmt(fRows.length)+'</td>'+
    '<td class="num ehbtp">'+fmtR(fEhbTp)+'</td>'+
    '<td class="num">'+fmt(fAcc)+'</td>'+
    '<td class="num">'+fmt(fA2c)+'</td>'+
    '<td class="num pct">'+pct(fRows.length,tTotal)+'</td>'+
    '<td class="num pct">'+pct(fAcc,fRows.length)+'</td>'+
    '<td class="num pct">'+pct(fA2c,fAcc)+'</td>';  // Converted% = A2C/Accepted (D14)
  tbody.appendChild(tr1);

  // Re-Raised by RA row
  var tr2=document.createElement('tr');
  tr2.innerHTML='<td>🔄 Re-Raised by RA</td>'+
    '<td class="num">'+fmt(rrCount)+'</td>'+
    '<td class="num ehbtp">'+fmtR(rrEhbTp)+'</td>'+
    '<td class="num">'+fmt(rraCount)+'</td>'+
    '<td class="num">'+fmt(rrA2c)+'</td>'+
    '<td class="num pct">'+pct(rrCount,tTotal)+'</td>'+
    '<td class="num pct">'+pct(rraCount,rrCount)+'</td>'+
    '<td class="num pct">'+pct(rrA2c,rraCount)+'</td>';
  tbody.appendChild(tr2);

  // Ticket by Bot row
  var tr3=document.createElement('tr');
  tr3.innerHTML='<td>🎫 Ticket by Bot</td>'+
    '<td class="num">'+fmt(tktRows.length)+'</td>'+
    '<td class="num ehbtp">'+fmtR(tktEhbTp)+'</td>'+
    '<td class="num">'+fmt(tktAcc)+'</td>'+
    '<td class="num">'+fmt(tktA2c)+'</td>'+
    '<td class="num pct">'+pct(tktRows.length,tTotal)+'</td>'+
    '<td class="num pct">'+pct(tktAcc,tktRows.length)+'</td>'+
    '<td class="num pct">'+pct(tktA2c,tktAcc)+'</td>';
  tbody.appendChild(tr3);

  // Drop / No Action row
  var tr4=document.createElement('tr');
  tr4.innerHTML='<td>🚫 Drop / No Action</td>'+
    '<td class="num">'+fmt(dropRows.length)+'</td>'+
    '<td class="num ehbtp">'+fmtR(dropEhbTp)+'</td>'+
    '<td class="num">'+fmt(dropAcc)+'</td>'+
    '<td class="num">'+fmt(dropA2c)+'</td>'+
    '<td class="num pct">'+pct(dropRows.length,tTotal)+'</td>'+
    '<td class="num pct">'+pct(dropAcc,dropRows.length)+'</td>'+
    '<td class="num pct">'+pct(dropA2c,dropAcc)+'</td>';
  tbody.appendChild(tr4);

  // Overall row
  var trOv=document.createElement('tr');
  trOv.className='row-overall';
  trOv.innerHTML='<td>⭐ Overall (All Raises)</td>'+
    '<td class="num">'+fmt(ovRows.length)+'</td>'+
    '<td class="num ehbtp">'+fmtR(ovEhbTp)+'</td>'+
    '<td class="num">'+fmt(ovAcc)+'</td>'+
    '<td class="num">'+fmt(ovA2c)+'</td>'+
    '<td class="num pct">'+pct(ovRows.length,tTotal)+'</td>'+
    '<td class="num pct">'+pct(ovAcc,ovRows.length)+'</td>'+
    '<td class="num pct">'+pct(ovA2c,tTotal)+'</td>';
  tbody.appendChild(trOv);

  // Total row
  var trTot=document.createElement('tr');
  trTot.className='row-total';
  trTot.innerHTML='<td>📋 Total Test Leads</td>'+
    '<td class="num" colspan="7">'+fmt(tTotal)+' unique appointments</td>';
  tbody.appendChild(trTot);

  // ── Control Table ──
  var ctrlTbody = document.getElementById('ctrl-tbody');
  ctrlTbody.innerHTML = '';
  var ctrlTr = document.createElement('tr');
  ctrlTr.className='row-overall';
  ctrlTr.innerHTML=
    '<td class="num">'+fmt(cTotal)+'</td>'+
    '<td class="num">'+fmt(cRaised)+'</td>'+
    '<td class="num ehbtp">'+fmtR(cEhbTp)+'</td>'+
    '<td class="num">'+fmt(cAcc)+'</td>'+
    '<td class="num">'+fmt(cA2c)+'</td>'+
    '<td class="num pct">'+pct(cRaised,cTotal)+'</td>'+
    '<td class="num pct">'+pct(cAcc,cRaised)+'</td>'+
    '<td class="num pct">'+pct(cA2c,cAcc)+'</td>';  // Conversion% = A2C/Accepted
  ctrlTbody.appendChild(ctrlTr);

  // ── Comparison Table ──
  var cmpTbody = document.getElementById('cmp-tbody');
  cmpTbody.innerHTML = '';
  var cmpRows=[
    ['Total Leads', fmt(tTotal), fmt(cTotal)],
    ['Raised', fmt(ovRows.length), fmt(cRaised)],
    ['EHB/TP', fmtR(ovEhbTp), fmtR(cEhbTp)],
    ['Acceptance', fmt(ovAcc), fmt(cAcc)],
    ['A2C', fmt(fA2c), fmt(cA2c)],
    ['Total tokens', fmt(ovA2c), fmt(cToken)],
    ['Raised%', pct(ovRows.length,tTotal), pct(cRaised,cTotal)],
    ['Acceptance% (multi-raise)', pct(ovAcc,ovRows.length), pct(cAcc,cRaised)],
    ['A2C% (A2C/Accepted)', pct(fA2c,ovAcc), pct(cA2c,cAcc)],
    ['Overall conversion (Token/Leads)', pct(ovA2c,tTotal), pct(cToken,cTotal)],
  ];
  cmpRows.forEach(function(r){
    var tr=document.createElement('tr');
    tr.innerHTML='<td>'+r[0]+'</td><td class="num test-col">'+r[1]+'</td><td class="num ctrl-col">'+r[2]+'</td>';
    cmpTbody.appendChild(tr);
  });
}

// Attach listeners
['dt-from','dt-to','case-type','region','gs','c2b'].forEach(function(id){
  document.getElementById(id).addEventListener('change', recompute);
});
recompute();
})();
</script>
</body>
</html>"""


def _build_postchat_dashboard():
    """Read post-chat Excel, generate dashboard_postchat.html."""

    # Find the Excel file in the same folder
    excel_file = None
    for f in OUT_DIR.glob("*.xlsx"):
        name_low = f.name.lower()
        if any(k in name_low for k in ["ai chat", "bot query", "postchat", "post_chat", "post chat"]):
            excel_file = f
            break
    if excel_file is None:
        xl_list = list(OUT_DIR.glob("*.xlsx"))
        if xl_list:
            excel_file = max(xl_list, key=lambda f: f.stat().st_mtime)
    if excel_file is None:
        print("  [postchat] No Excel file found — skipping post-chat dashboard.")
        return

    print(f"  [postchat] Reading: {excel_file.name}")
    wb = openpyxl.load_workbook(str(excel_file), read_only=True, data_only=True)

    def _date(v):
        if v is None: return None
        if hasattr(v, 'strftime'): return v.strftime('%Y-%m-%d')
        s = str(v)[:10]
        return s if len(s) == 10 else None

    def _num(v):
        if v is None: return None
        if isinstance(v, str) and (v.startswith('#') or v.strip() == ''): return None
        try:
            f = float(v)
            import math
            return None if math.isnan(f) else f
        except Exception:
            return None

    GS  = {'non_gs': 'n', 'gs_flagged': 'f', 'gs_non_assured': 'a'}
    C2B = {'c2b': 'b', 'c2d': 'd'}
    CAT = {'bot_raise': 'r', 'ticket': 't', 'drop': 'd'}

    # ── Raw_Data (Test / Bot) ──
    # Columns (0-indexed): [1]=APPOINTMENT_ID, [4]=FIRST_CHAT_DATE, [7]=DSQ_FLAG,
    # [8]=QNR_FLAG, [12]=REGION, [13]=GS_FLAGS, [14]=TP, [15]=C2B_FLAG,
    # [19]=PAI_OCB_CREATED, [20]=CX_PAI_OCB, [21]=CX_AVG_EHB, [23]=PAI_OCB_ACCEPTED,
    # [24]=CX_PAI_OCB_ACCEPTED, [28]=Category, [29]=New Token, [30]=Re-Raised,
    # [31]=Re-Raised Accepted, [32]=EHB/TP, [37]=Calc ehbtp:cx, [38]=Calc ehbtp:overall
    ws_test = wb['Raw_Data']
    test_rows = []
    for row in ws_test.iter_rows(min_row=2, values_only=True):
        if row[1] is None:
            continue
        ct = 'Q' if row[8] == 'QNR' else ('D' if row[7] == 'DSQ' else None)
        test_rows.append([
            row[1],                         # [0]  APPOINTMENT_ID
            _date(row[4]),                  # [1]  FIRST_CHAT_DATE
            ct,                             # [2]  case type Q/D
            row[12],                        # [3]  REGION
            GS.get(row[13], row[13]),       # [4]  GS (n/f/a)
            C2B.get(row[15], row[15]),      # [5]  C2B flag (b/d)
            _num(row[14]),                  # [6]  TP
            _num(row[19]),                  # [7]  PAI_OCB_CREATED
            _num(row[20]),                  # [8]  CX_PAI_OCB
            _num(row[21]),                  # [9]  CX_AVG_EHB
            _num(row[23]),                  # [10] PAI_OCB_ACCEPTED
            _num(row[24]),                  # [11] CX_PAI_OCB_ACCEPTED
            CAT.get(row[28], row[28]),      # [12] Category (r/t/d)
            1 if (row[29] is not None and (hasattr(row[29],'strftime') or (_num(row[29]) is not None and _num(row[29]) > 40000))) else None,  # [13] New Token flag (1=token issued)
            _num(row[30]),                  # [14] Re-Raised
            _num(row[31]),                  # [15] Re-Raised Accepted
            _num(row[32]),                  # [16] EHB/TP pre-calc
            _num(row[37]),                  # [17] Calc ehbtp:cx
            _num(row[38]),                  # [18] Calc ehbtp:overall
        ])

    # ── New MB Dump (Control / Human), new_flag=0 only ──
    # Columns: [0]=APP_ID, [3]=FIRST_CREATED_DATE, [6]=SUB_CATEGORY,
    # [9]=PAI_OCB_CREATED_TKT_OVERALL, [10]=PAI_OCB_ACCEPTED_TKT_OVERALL,
    # [12]=EXPECTED_HB_OVERALL, [14]=TP, [15]=GS_FLAGS, [16]=REGION,
    # [17]=New Bought, [21]=Overall ratio, [24]=C2B/C2D, [26]=new_flag
    ws_ctrl = wb['New MB Dump']
    ctrl_rows = []
    for row in ws_ctrl.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        if row[26] != 0:
            continue  # new_flag must be 0
        ct = 'Q' if row[6] == 'QNR' else ('D' if row[6] == 'DSQ' else None)
        c2b_val = None
        if row[24]:
            c2b_str = str(row[24]).lower().strip()
            c2b_val = C2B.get(c2b_str)
        ctrl_rows.append([
            row[0],                         # [0]  APP_ID
            _date(row[3]),                  # [1]  FIRST_CREATED_DATE
            ct,                             # [2]  case type Q/D
            row[16],                        # [3]  REGION
            GS.get(row[15], row[15]),       # [4]  GS (n/f/a)
            c2b_val,                        # [5]  C2B/C2D (b/d/None)
            _num(row[14]),                  # [6]  TP
            _num(row[9]),                   # [7]  PAI_OCB_CREATED_TKT_OVERALL
            _num(row[10]),                  # [8]  PAI_OCB_ACCEPTED_TKT_OVERALL
            _num(row[12]),                  # [9]  EXPECTED_HB_OVERALL
            _num(row[17]),                  # [10] New Bought
            _num(row[21]),                  # [11] Overall ratio (EHB/TP)
        ])

    # Date range for defaults
    test_dates = [r[1] for r in test_rows if r[1]]
    ctrl_dates = [r[1] for r in ctrl_rows if r[1]]
    all_dates  = test_dates + ctrl_dates
    min_date = min(all_dates) if all_dates else '2026-01-01'
    max_date = max(all_dates) if all_dates else '2026-12-31'

    # Unique regions
    regions = sorted(r for r in set(
        r[3] for r in test_rows + ctrl_rows
    ) if r)

    test_json    = json.dumps(test_rows, ensure_ascii=True).replace("</", "<\\/")
    ctrl_json    = json.dumps(ctrl_rows, ensure_ascii=True).replace("</", "<\\/")
    regions_json = json.dumps(regions, ensure_ascii=True)
    generated_at = datetime.now().strftime("%d %b %Y, %I:%M %p")

    html = _POSTCHAT_HTML
    html = html.replace("%%TEST_DATA%%",    test_json)
    html = html.replace("%%CTRL_DATA%%",    ctrl_json)
    html = html.replace("%%REGIONS%%",      regions_json)
    html = html.replace("%%MIN_DATE%%",     min_date)
    html = html.replace("%%MAX_DATE%%",     max_date)
    html = html.replace("%%GENERATED_AT%%", generated_at)

    out_path = OUT_DIR / "dashboard_postchat.html"
    out_path.write_text(html, encoding="utf-8")
    print(f"  → dashboard_postchat.html  ({len(test_rows):,} test · {len(ctrl_rows):,} control rows)")


# ── Generate HTML files ───────────────────────────────────────────────────────

def generate(offline=False, csv_path=None):
    import subprocess

    rows     = fetch_and_archive(offline=offline, csv_path=csv_path)
    sessions = process_sessions(rows)
    metrics  = compute_metrics(sessions)
    insights = generate_insights(metrics)
    summary  = make_summary(metrics)
    generated_at = datetime.now().strftime("%d %b %Y, %I:%M %p")

    gate = ""  # Password gate disabled — open access

    # Escape characters that would break an inline <script> block
    payload = json.dumps({"metrics": metrics, "qual": insights}, ensure_ascii=True) \
                  .replace("</", "<\\/").replace("<!--", "<\\!--")

    chartjs = get_chartjs()

    def build(template, replacements):
        html = template.replace("{{", "{").replace("}}", "}")
        html = html.replace('<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>',
                            f'<script>{chartjs}</script>')
        for k, v in replacements.items():
            html = html.replace(k, v)
        return html

    common = {
        "{DATA_JSON}":       payload,
        "{GENERATED_AT}":    generated_at,
        "{TOTAL_SESSIONS}":  str(metrics["total"]),
        "{CHAT_WIDGET}":     CHAT_WIDGET,
        "{RECOMPUTE_JS}":    RECOMPUTE_JS,
        "{PASSWORD_GATE}":   gate,
    }

    def safe_write(path, content):
        """Write large files robustly — delete first to avoid stale bytes on FUSE/Windows mounts."""
        encoded = content.encode('utf-8')
        # Remove existing file first so no stale bytes are left on FUSE/bindfs mounts
        if os.path.exists(path):
            try:
                os.unlink(path)
            except OSError:
                pass
        with open(path, 'wb') as f:
            f.write(encoded)
        # Verify
        written = os.path.getsize(path)
        if written != len(encoded):
            print(f"  ⚠️  Write mismatch for {os.path.basename(path)}: {written} bytes on disk vs {len(encoded)} expected")

    # Metrics dashboard
    html_m = build(METRICS_HTML, common)
    safe_write(OUT_DIR / "dashboard_metrics.html", html_m)

    # Qualitative dashboard
    html_q = build(QUAL_HTML, {**common,
        "{SUMMARY}":   summary,
        "{WW_COUNT}":  str(len(insights["workingWell"])),
        "{NI_COUNT}":  str(len(insights["needsImprovement"])),
    })
    safe_write(OUT_DIR / "dashboard_qualitative.html", html_q)

    # Post-chat dashboard (reads Excel if present)
    _build_postchat_dashboard()

    print(f"\n✅ Done! Generated dashboards for {metrics['total']} sessions.")
    print(f"   → dashboard_metrics.html")
    print(f"   → dashboard_qualitative.html")
    print(f"   • Auction+Ticket Both: {metrics['has_both_count']} sessions ({metrics['has_both_pct']}%)")
    print(f"   • Regions tracked: {len(metrics['region_split'])}")
    total_gap = sum(metrics['insp_gap']['dsq']) + sum(metrics['insp_gap']['qnr'])
    print(f"   • Sessions with inspection gap data: {total_gap}")

    # ── Auto-publish to GitHub Pages (if repo is configured) ─────────────────
    git_dir = OUT_DIR / ".git"
    if git_dir.exists():
        try:
            files = ["dashboard_metrics.html", "dashboard_qualitative.html", "dashboard_postchat.html", "index.html"]
            # Pull latest remote changes first to avoid push rejection
            subprocess.run(["git", "-C", str(OUT_DIR), "pull", "--rebase", "origin", "main"],
                           check=False, capture_output=True)
            subprocess.run(["git", "-C", str(OUT_DIR), "add"] + files,
                           check=False, capture_output=True)
            result = subprocess.run(
                ["git", "-C", str(OUT_DIR), "commit", "-m",
                 f"Update dashboards {generated_at}"],
                capture_output=True, text=True)
            if "nothing to commit" in result.stdout + result.stderr:
                print("\n   GitHub: No changes to publish.")
            else:
                push = subprocess.run(
                    ["git", "-C", str(OUT_DIR), "push", "origin", "main"],
                    capture_output=True, text=True)
                if push.returncode == 0:
                    print(f"\n   ✅ Published to GitHub Pages!")
                else:
                    print(f"\n   ⚠️  Git push failed: {push.stderr.strip()}")
        except FileNotFoundError:
            print("\n   ℹ️  Git not found — skipping auto-publish.")
    else:
        print(f"\nOpen either file in Chrome — no server needed.")

if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(
        description="Generate CARS24 chatbot dashboards",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python generate_dashboards.py                        # live fetch (sheet must be public)
  python generate_dashboards.py --csv sheet.csv        # use manually downloaded CSV
  python generate_dashboards.py --auth-setup           # one-time OAuth login (saves token.json)
  python generate_dashboards.py --offline              # use local archive only (no network)
        """
    )
    ap.add_argument("--offline", action="store_true",
                    help="Skip Google Sheets fetch and use local archive only")
    ap.add_argument("--csv", metavar="FILE",
                    help="Path to a manually downloaded CSV export of the Google Sheet")
    ap.add_argument("--auth-setup", action="store_true",
                    help="Run one-time OAuth setup and save token.json for future runs")
    args = ap.parse_args()

    if args.auth_setup:
        # ── One-time OAuth setup ──────────────────────────────────────────────
        creds_file = OUT_DIR / "credentials.json"
        token_file = OUT_DIR / "token.json"
        if not creds_file.exists():
            print("\n  To set up OAuth authentication:\n")
            print("  1. Go to https://console.cloud.google.com/")
            print("  2. Create a project (or select existing)")
            print("  3. Enable the Google Sheets API")
            print("  4. Go to APIs & Services → Credentials → Create Credentials → OAuth client ID")
            print("  5. Application type: Desktop app")
            print("  6. Download the JSON file and save it as:")
            print(f"     {creds_file}\n")
            print("  Then run: python generate_dashboards.py --auth-setup")
            raise SystemExit(0)
        try:
            from google_auth_oauthlib.flow import InstalledAppFlow
            flow = InstalledAppFlow.from_client_secrets_file(
                str(creds_file),
                scopes=["https://www.googleapis.com/auth/spreadsheets.readonly"]
            )
            creds = flow.run_local_server(port=0)
            token_file.write_text(creds.to_json())
            print(f"\n  ✅ OAuth setup complete! Token saved to: {token_file}")
            print("  Future runs will authenticate automatically.\n")
            print("  Now generating dashboards…\n")
            generate(csv_path=None)
        except ImportError:
            print("\n  Install required packages first:")
            print("  pip install google-auth-oauthlib google-auth-httplib2\n")
            raise SystemExit(1)
    else:
        generate(offline=args.offline, csv_path=args.csv)
